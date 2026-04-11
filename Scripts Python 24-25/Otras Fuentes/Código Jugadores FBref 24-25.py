"""
Código Jugadores FBref 24-25.py
================================
Versión de la temporada pasada (LaLiga 2024-25) del script Código Jugadores FBref.py.
Parsea los HTML de FBref y genera Jugadores FBref 24-25.xlsx en Datos/Temporada Pasada/.

Ver Código Jugadores FBref.py (temporada actual) para documentación completa.

Uso: python3 "Temporada Pasada/Código Jugadores FBref 24-25.py"
"""

from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ──────────────────────────────────────────────────────────────────────
DATA_DIR  = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Descargas FBref/Temporada Pasada"
OUTPUT    = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Temporada Pasada/Jugadores FBref 24-25.xlsx"

FILES = [
    ("Stats",      "Jugadores_Stats24-25.html",      "stats_standard"),
    ("Porteros",   "Jugadores_Porteros24-25.html",   "stats_keeper"),
    ("Tiros",      "Jugadores_Tiros24-25.html",       "stats_shooting"),
    ("Tiempo",     "Jugadores_Tiempo24-25.html",      "stats_playing_time"),
    ("Miscelaneo", "Jugadores_Miscelaneo24-25.html",  "stats_misc"),
]

KEY_COLS       = ["player", "team", "birth_year"]
EXCLUDE_COLS   = {
    "ranker", "matches",
    "age", "goals_assists", "goals_assists_per90", "goals_assists_pens_per90",
    "pens_won", "pens_conceded",
}
BASE_INFO_COLS = {"player", "nationality", "position", "team", "age", "birth_year"}

# Columnas cuyos valores son porcentajes en el HTML (sin símbolo %)
PCT_COLS = {
    "shots_on_target_pct", "minutes_pct",
    "gk_save_pct", "gk_clean_sheets_pct", "gk_pens_save_pct",
}

RENAME = {
    "player": "Jugador", "nationality": "Nacionalidad", "position": "Posición",
    "team": "Equipo", "birth_year": "Año Nac.",
    "games": "PJ", "games_starts": "Titular", "minutes": "Min",
    "minutes_90s": "90s", "goals": "Goles", "assists": "Asistencias",
    "goals_pens": "G-PK", "pens_made": "PK",
    "pens_att": "PKint", "cards_yellow": "Amarillas", "cards_red": "Rojas",
    "goals_per90": "Gls/90", "assists_per90": "Ast/90",
    "goals_pens_per90": "G-PK/90",
    # Porteros
    "gk_games": "PJ", "gk_games_starts": "Titular",
    "gk_minutes": "Min", "gk_goals_against": "GA",
    "gk_goals_against_per90": "GA90", "gk_shots_on_target_against": "aTiro",
    "gk_saves": "Paradas", "gk_save_pct": "% Paradas",
    "gk_wins": "W", "gk_ties": "E", "gk_losses": "L",
    "gk_clean_sheets": "Portería0", "gk_clean_sheets_pct": "% P0",
    "gk_pens_att": "PK rec", "gk_pens_allowed": "PK concedidos",
    "gk_pens_saved": "PK parados", "gk_pens_missed": "PK fallados",
    "gk_pens_save_pct": "% PK parados",
    # Tiros
    "shots": "Tiros", "shots_on_target": "Tiros a puerta",
    "shots_on_target_pct": "% Tiros puerta", "shots_per90": "Tiros/90",
    "shots_on_target_per90": "Tiros P/90", "goals_per_shot": "Gls/Tiro",
    "goals_per_shot_on_target": "Gls/Tiro P",
    # Tiempo de juego
    "minutes_per_game": "Min/PJ", "minutes_pct": "% Min",
    "minutes_per_start": "Min/Tit", "games_complete": "Completos",
    "games_subs": "Suplente", "minutes_per_sub": "Min/Sub",
    "unused_subs": "Sub no usado", "points_per_game": "Pts/PJ",
    "on_goals_for": "GF con él", "on_goals_against": "GC con él",
    "plus_minus": "+/-", "plus_minus_per90": "+/-90",
    "plus_minus_wowy": "+/- On-Off",
    # Misceláneo
    "cards_yellow_red": "2ªAmar", "fouls": "Faltas com",
    "fouled": "Faltas rec", "offsides": "Fuera de juego",
    "crosses": "Centros", "interceptions": "Intercep",
    "tackles_won": "Entradas gan", "own_goals": "Autogoles",
}

POS_MAP = {
    "GK": "Portero",
    "DF": "Defensa",
    "MF": "Centrocampista",
    "FW": "Delantero",
}

TEXT_COLS = {"player", "nationality", "position", "team", "birth_year",
             "Jugador", "Nacionalidad", "Posición", "Equipo", "Año Nac."}


# ── Transformaciones ───────────────────────────────────────────────────────────
def clean_nation(value: str) -> str:
    return " ".join(re.findall(r'[A-Z]+', value))

def clean_position(value: str) -> str:
    parts = [p.strip() for p in value.split(",")]
    return ", ".join(POS_MAP.get(p, p) for p in parts if p)

def to_number(value: str):
    """Convierte string a float, eliminando el separador de miles (coma anglosajona)."""
    if value == "" or value is None:
        return None
    try:
        return float(value.replace(",", ""))
    except (ValueError, TypeError):
        return value

def apply_transformations(df: pd.DataFrame) -> pd.DataFrame:
    if "nationality" in df.columns:
        df["nationality"] = df["nationality"].apply(clean_nation)
    if "position" in df.columns:
        df["position"] = df["position"].apply(clean_position)
    return df

def apply_numeric_conversion(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if col not in TEXT_COLS:
            df[col] = df[col].apply(lambda v: to_number(str(v)) if v != "" else None)
    return df


# ── Parseo HTML ────────────────────────────────────────────────────────────────
def parse_table(filename: str, table_id: str) -> pd.DataFrame:
    path = f"{DATA_DIR}/{filename}"
    with open(path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    table = soup.find("table", id=table_id)
    thead = table.find("thead")
    col_stats = [th.get("data-stat", "") for th in thead.find_all("tr")[-1].find_all("th")]

    rows = []
    for tr in table.find("tbody").find_all("tr"):
        if tr.get("class") and "thead" in tr.get("class"):
            continue
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        row = {stat: cell.get_text(strip=True)
               for stat, cell in zip(col_stats, cells)
               if stat and stat not in EXCLUDE_COLS}
        rows.append(row)

    df = pd.DataFrame(rows)
    if "player" in df.columns:
        df = df[df["player"].notna() & (df["player"] != "")]
    return df


# ── Construcción de DataFrames ─────────────────────────────────────────────────
def build_dataframes():
    raw = {name: parse_table(filename, table_id) for name, filename, table_id in FILES}
    portero_names = set(raw["Porteros"]["player"].tolist())

    # ── Hoja Jugadores ──────────────────────────────────────────────────────
    merged = raw["Stats"][~raw["Stats"]["player"].isin(portero_names)].copy()
    for name in ["Tiros", "Tiempo", "Miscelaneo"]:
        df = raw[name][~raw[name]["player"].isin(portero_names)].copy()
        drop_cols = [c for c in df.columns if c in BASE_INFO_COLS and c not in KEY_COLS]
        df = df.drop(columns=drop_cols, errors="ignore")
        dup_cols = [c for c in df.columns if c in set(merged.columns) and c not in KEY_COLS]
        df = df.drop(columns=dup_cols, errors="ignore")
        merged = pd.merge(merged, df, on=KEY_COLS, how="outer")
    merged = merged.fillna("")
    merged = apply_transformations(merged)
    merged = merged[merged["nationality"] != ""]
    merged = merged.rename(columns={k: v for k, v in RENAME.items() if k in merged.columns})
    df_field = apply_numeric_conversion(merged)

    # ── Hoja Porteros ───────────────────────────────────────────────────────
    df_gk = raw["Porteros"].copy().fillna("")
    df_gk = apply_transformations(df_gk)
    df_gk = df_gk[df_gk["nationality"] != ""]
    df_gk = df_gk.rename(columns={k: v for k, v in RENAME.items() if k in df_gk.columns})
    df_gk = apply_numeric_conversion(df_gk)

    return df_field, df_gk


# ── Columnas de porcentaje (nombres ya renombrados) ────────────────────────────
PCT_COLS_RENAMED = {RENAME.get(c, c) for c in PCT_COLS}


# ── Escritura y estilo Excel ───────────────────────────────────────────────────
def write_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str):
    ws = wb.create_sheet(title=sheet_name)

    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
    NORMAL_FILL = PatternFill("solid", start_color="FFFFFF")
    BORDER_SIDE = Side(style="thin", color="BFBFBF")
    CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                         top=BORDER_SIDE, bottom=BORDER_SIDE)

    # Determinar formato por columna
    col_formats = {}
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in TEXT_COLS:
            col_formats[col_idx] = "@"
        elif col_name in PCT_COLS_RENAMED:
            col_formats[col_idx] = '#,##0.0"%"'
        else:
            sample_nums = [v for v in df[col_name].dropna() if isinstance(v, float)]
            has_decimals = any(v != int(v) for v in sample_nums[:20]) if sample_nums else False
            col_formats[col_idx] = "#,##0.00" if has_decimals else "#,##0"

    # Cabecera
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 30

    # Datos
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else NORMAL_FILL
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value if value not in ("", None) else None
            cell.number_format = col_formats[col_idx]
            cell.fill = fill
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER

    # Ancho de columnas
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), 6)
        for row_idx in range(2, min(len(df) + 2, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 22)

    ws.freeze_panes = "B2"


# ── Main ───────────────────────────────────────────────────────────────────────
df_field, df_gk = build_dataframes()

wb = Workbook()
wb.remove(wb.active)
write_sheet(wb, df_field, "Jugadores")
write_sheet(wb, df_gk, "Porteros")
wb.save(OUTPUT)

print(f"✅ Excel creado: {OUTPUT}")
print(f"   Hoja Jugadores: {len(df_field)} jugadores | {len(df_field.columns)} columnas")
print(f"   Hoja Porteros:  {len(df_gk)} porteros  | {len(df_gk.columns)} columnas")
