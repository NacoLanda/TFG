"""
Código Jugadores Unificados 24-25.py
======================================
Versión de la temporada pasada (LaLiga 2024-25) del script Código Jugadores Unificados.py.
Cruza FBref con WhoScored y genera Jugadores Unificados 24-25.xlsx en Datos/Temporada Pasada/.

Ver Código Jugadores Unificados.py (temporada actual) para documentación completa.

Uso: python3 "Temporada Pasada/Código Jugadores Unificados 24-25.py"
"""

import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RUTA      = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Temporada Pasada/"
FBREF     = RUTA + "Jugadores FBref 24-25.xlsx"
WHOSCORED = RUTA + "Datos WhoScored 24-25.xlsx"
SALIDA    = RUTA + "Jugadores Unificados 24-25.xlsx"
UMBRAL    = 85


# Correcciones manuales de nombres FBref → WhoScored para la temporada 2024-25.
# Añade aquí los casos que el fuzzy matching no resuelva correctamente.
# Formato: "Nombre en FBref": "Nombre en WhoScored"  (o None para excluir)
CORRECCIONES = {
}


# Genera la lista de triplets (col_general, col_local, col_visitante) para una
# lista de nombres base de WhoScored.
def _ws(*bases):
    cols = []
    for b in bases:
        cols += [f"{b}_general", f"{b}_local", f"{b}_visitante"]
    return cols

CATEGORIAS_JUG = [
    ("IDENTIFICACIÓN", ["Jugador","Nacionalidad","Posición","Equipo","Año Nac."]),

    # ── PRESENCIA ──────────────────────────────────────────────────────────────
    ("PRESENCIA", [
        "PJ","Titular","Min","90s","Completos",
        "Min/PJ","% Min","Min/Tit","Min/Sub","Sub no usado","Pts/PJ",
        *_ws("jgdos","mins","rating","suplente"),
    ]),

    # ── ATAQUE ─────────────────────────────────────────────────────────────────
    ("ATAQUE", [
        "Goles","Asistencias","G-PK","PK","Gls/90","Ast/90","G-PK/90",
        "Tiros","Tiros a puerta","% Tiros puerta","Tiros/90","Tiros P/90",
        "Gls/Tiro","Gls/Tiro P","Autogoles",
        *_ws("goles","asist"),
        *_ws("tpp","tiros"),
        *_ws("xg","xg_dif","xg_90","xg_tiros"),
        *_ws("propia","pclave"),
        *_ws("tir_z_FueraArea","tir_z_AreaPeq","tir_z_AreaPenalti"),
        *_ws("tir_s_JuegoAb","tir_s_Contra","tir_s_BParado","tir_s_Penaltis"),
        *_ws("tir_p_FueraPorteria","tir_p_AlPoste","tir_p_APorteria","tir_p_Bloqueado"),
        *_ws("tir_c_PieDer","tir_c_PieIzq","tir_c_Cabeza","tir_c_Otro"),
        *_ws("gol_z_AreaPeq","gol_z_AreaPenalti","gol_z_FueraArea"),
        *_ws("gol_s_JuegoAb","gol_s_Contra","gol_s_BParado","gol_s_Penaltis","gol_s_Normal"),
        *_ws("gol_c_PieDer","gol_c_PieIzq","gol_c_Cabeza","gol_c_Otro"),
        *_ws("ast_Centro","ast_Corner","ast_PHueco","ast_TiroLibr","ast_SBanda","ast_Otro"),
    ]),

    # ── PASES ──────────────────────────────────────────────────────────────────
    ("PASES", [
        *_ws("ap_pct","prome_p","blargos","phueco","centr","pdasb"),
        *_ws("pas_l_BLPrec","pas_l_BLImp","pas_l_PCortoPre","pas_l_PCortoImp"),
        *_ws("pas_t_CentrPrec","pas_t_CentrImp",
             "pas_t_CrnPrec","pas_t_CrnImp",
             "pas_t_TirLibPrec","pas_t_TirLibImp"),
        *_ws("pc_l_Largo","pc_l_Corto"),
        *_ws("pc_t_Centro","pc_t_Corner","pc_t_PHueco",
             "pc_t_TiroLibr","pc_t_SBanda","pc_t_Otro"),
    ]),

    # ── DEFENSA ────────────────────────────────────────────────────────────────
    ("DEFENSA", [
        "Intercep","Entradas gan","Centros",
        *_ws("entrad","interc","despe","bloq","aereos"),
        *_ws("bloq2_TirosParados","bloq2_CentrBloq","bloq2_PasesBloq"),
        *_ws("aer_Perdidos"),
        *_ws("par_AreaPeq","par_AreaPenalti","par_FueraArea"),
    ]),

    # ── DISCIPLINA ─────────────────────────────────────────────────────────────
    ("DISCIPLINA", [
        "Amarillas","Rojas","2ªAmar","Faltas com","Faltas rec","PKint",
        *_ws("amar","roja","falt","fjuego_g"),
    ]),

    # ── RENDIMIENTO ────────────────────────────────────────────────────────────
    ("RENDIMIENTO", [
        "GF con él","GC con él","+/-","+/-90","+/- On-Off","Fuera de juego",
    ]),

    # ── DUELOS Y PÉRDIDAS ──────────────────────────────────────────────────────
    ("DUELOS Y PÉRDIDAS", [
        *_ws("rgts","rgt_NoExitoso","despo"),
    ]),

    # ── TÁCTICA OFENSIVA ───────────────────────────────────────────────────────
    ("TÁCTICA OFENSIVA", [
        *_ws("jdelp","faltf","fjuego_c"),
    ]),

    # ── NUEVAS ─────────────────────────────────────────────────────────────────
    ("NUEVAS", [
        *_ws("rgt_Exitoso"),
    ]),
]

CATEGORIAS_POR = [
    ("IDENTIFICACIÓN", ["Jugador","Nacionalidad","Posición","Equipo","Año Nac."]),

    ("PRESENCIA", [
        "PJ","Titular","Min","90s",
        *_ws("jgdos","mins","rating","suplente"),
    ]),

    ("RENDIMIENTO", [
        "GA","GA90","aTiro","Paradas","% Paradas","W","E","L","Portería0","% P0",
        *_ws("par_AreaPeq","par_AreaPenalti","par_FueraArea"),
    ]),

    ("PENALTIS", [
        "PK rec","PK concedidos","PK parados","PK fallados","% PK parados",
    ]),

    ("PASES", [
        *_ws("ap_pct","prome_p","blargos"),
    ]),

    ("DISCIPLINA", [
        *_ws("amar","roja","falt","fjuego_g"),
    ]),
]

FONT_NAME  = "Aptos Narrow"
AZUL_OSC   = "1F4E79"
AZUL       = "2D75B8"
AZUL_CLAR  = "BDD7EE"
BLANCO     = "FFFFFF"
NEGRO      = "000000"

CAT_COLORS = {
    "IDENTIFICACIÓN":    ("1F4E79", BLANCO),
    "PRESENCIA":         ("2D75B8", BLANCO),
    "ATAQUE":            ("C55A11", BLANCO),
    "PASES":             ("538135", BLANCO),
    "DEFENSA":           ("7030A0", BLANCO),
    "DISCIPLINA":        ("833C00", BLANCO),
    "RENDIMIENTO":       ("1F4E79", BLANCO),
    "DUELOS Y PÉRDIDAS": ("4472C4", BLANCO),
    "TÁCTICA OFENSIVA":  ("2E75B6", BLANCO),
    "PENALTIS":          ("BF9000", BLANCO),
}

def nombre_cabecera(col):
    if " | " in col:
        col = col.split(" | ", 1)[1]
    if col.endswith(".1") or col.endswith(".2"):
        col = col[:-2]
    for suf in ("_general", "_local", "_visitante"):
        if col.endswith(suf):
            col = col[:-len(suf)]
            break
    return col

def subcabecera(col):
    if col.endswith("_general"):    return "General"
    if col.endswith("_local"):      return "Local"
    if col.endswith("_visitante"):  return "Visitante"
    if " | " in col:
        if col.endswith(".1"):   return "Local"
        elif col.endswith(".2"): return "Visitante"
        else:                    return "General"
    return ""

def estilo(cell, valor, bg, fg, bold, size, wrap=False):
    cell.value     = valor
    cell.font      = Font(name=FONT_NAME, size=size, bold=bold, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def leer_whoscored(ruta, hoja):
    df = pd.read_excel(ruta, sheet_name=hoja, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    df.columns.values[0] = "JUGADOR"
    df = df.reset_index(drop=True)
    return df

def normalizar(n):
    return "" if pd.isna(n) else str(n).strip().lower()

def mejor_match(nombre, lista):
    n = normalizar(nombre)
    if not n: return None
    ln = [normalizar(x) for x in lista]
    res = process.extractOne(n, ln, scorer=fuzz.token_sort_ratio)
    if res and res[1] >= UMBRAL:
        return lista[ln.index(res[0])]
    return None

def unir(df_fb, df_ws):
    col_fb = df_fb.columns[0]
    if "Equipo" in df_fb.columns:
        df_fb = df_fb.drop_duplicates(subset=[col_fb, "Equipo"]).reset_index(drop=True)
    else:
        df_fb = df_fb.drop_duplicates(subset=[col_fb]).reset_index(drop=True)
    nombres_ws  = df_ws.iloc[:, 0].values.tolist()
    disponibles = nombres_ws.copy()
    match_dict  = {}
    for nb in df_fb[col_fb]:
        if nb in match_dict: continue
        if nb in CORRECCIONES:
            ws_n = CORRECCIONES[nb]
            if ws_n is None:
                match_dict[nb] = None
            elif ws_n in disponibles:
                match_dict[nb] = ws_n
                disponibles.remove(ws_n)
        else:
            m = mejor_match(nb, disponibles)
            if m:
                match_dict[nb] = m
                disponibles.remove(m)
    df_fb = df_fb.copy()
    df_fb["__key__"] = df_fb[col_fb].map(match_dict)
    df_ws = df_ws.copy()
    df_ws["__key__"] = df_ws.iloc[:, 0]
    df = df_fb.merge(df_ws, on="__key__", how="inner")
    df = df.drop(columns=["__key__", col_fb])
    df = df.rename(columns={"JUGADOR": "Jugador"})
    return df.reset_index(drop=True)

def escribir_hoja(wb, nombre_hoja, df, categorias):
    ws = wb.create_sheet(title=nombre_hoja)

    col_info = []
    vistas   = set()
    for cat, cols in categorias:
        for c in cols:
            if c in df.columns and c not in vistas:
                col_info.append((c, cat, subcabecera(c), nombre_cabecera(c)))
                vistas.add(c)
    for c in df.columns:
        if c not in vistas:
            col_info.append((c, "OTROS", "", c))

    n_cols = len(col_info)

    # Fila 1: categorías fusionadas
    ws.row_dimensions[1].height = 22
    cat_ini = 1
    cat_act = col_info[0][1]
    for idx in range(2, n_cols + 2):
        cat_now = col_info[idx - 1][1] if idx <= n_cols else None
        if cat_now != cat_act:
            bg, fg = CAT_COLORS.get(cat_act, (AZUL, BLANCO))
            cell = ws.cell(row=1, column=cat_ini)
            estilo(cell, cat_act, bg, fg, True, 12)
            if cat_ini < idx - 1:
                ws.merge_cells(f"{get_column_letter(cat_ini)}1:{get_column_letter(idx-1)}1")
            cat_act = cat_now
            cat_ini = idx

    # Fila 2: nombres de columnas
    ws.row_dimensions[2].height = 30
    for idx, (_, cat, _, nombre) in enumerate(col_info, start=1):
        cell = ws.cell(row=2, column=idx)
        if cat == "IDENTIFICACIÓN":
            estilo(cell, nombre, AZUL_OSC, BLANCO, True, 11, wrap=True)
        else:
            estilo(cell, nombre, AZUL_CLAR, NEGRO, True, 10, wrap=True)

    # Fila 3: General / Local / Visitante
    ws.row_dimensions[3].height = 16
    for idx, (_, cat, subcab, _) in enumerate(col_info, start=1):
        cell = ws.cell(row=3, column=idx)
        if subcab:
            estilo(cell, subcab, AZUL_CLAR, NEGRO, False, 9)
        else:
            bg, fg = CAT_COLORS.get(cat, (AZUL, BLANCO))
            estilo(cell, "", bg, fg, False, 9)

    # Datos desde fila 4
    cols_orden = [ci[0] for ci in col_info]
    for row_idx, row in enumerate(df[cols_orden].itertuples(index=False), start=4):
        ws.row_dimensions[row_idx].height = 15
        for col_idx, val in enumerate(row, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.value     = None if (isinstance(val, float) and pd.isna(val)) else val
            cell.font      = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Anchos
    for idx, (_, cat, _, _) in enumerate(col_info, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 16 if cat == "IDENTIFICACIÓN" else 9

fb_jug = pd.read_excel(FBREF, sheet_name="Jugadores")
fb_por = pd.read_excel(FBREF, sheet_name="Porteros")
ws_todos = leer_whoscored(WHOSCORED, "Jugadores")
# Resolver nombres de porteros FBref a sus equivalentes en WhoScored usando
# CORRECCIONES y fuzzy matching, igual que hace unir()
nombres_ws_todos = ws_todos["JUGADOR"].tolist()
nombres_porteros_ws = set()
for nb in fb_por["Jugador"].dropna().str.strip():
    if nb in CORRECCIONES:
        ws_n = CORRECCIONES[nb]
        if ws_n and ws_n in nombres_ws_todos:
            nombres_porteros_ws.add(ws_n)
    elif nb in nombres_ws_todos:
        nombres_porteros_ws.add(nb)
    else:
        m = mejor_match(nb, nombres_ws_todos)
        if m:
            nombres_porteros_ws.add(m)
ws_por = ws_todos[ws_todos["JUGADOR"].isin(nombres_porteros_ws)].reset_index(drop=True)
ws_jug = ws_todos[~ws_todos["JUGADOR"].isin(nombres_porteros_ws)].reset_index(drop=True)

df_jugadores = unir(fb_jug, ws_jug)
df_porteros  = unir(fb_por, ws_por)

wb = Workbook()
wb.remove(wb.active)
escribir_hoja(wb, "Jugadores", df_jugadores, CATEGORIAS_JUG)
escribir_hoja(wb, "Porteros",  df_porteros,  CATEGORIAS_POR)
wb.save(SALIDA)

print(f"Jugadores: {len(df_jugadores)} filas")
print(f"Porteros:  {len(df_porteros)} filas")
print("Guardado en:", SALIDA)
