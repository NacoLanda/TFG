"""
Código FBref 24-25.py
======================
Versión de la temporada pasada (LaLiga 2024-25) del script Código FBref.py.
Parsea el HTML descargado de FBref y genera Datos FBref.xlsx en la carpeta
Datos/Temporada Pasada/.

Ver Código FBref.py (temporada actual) para documentación completa.

Uso: python3 "Temporada Pasada/Código FBref 24-25.py"
"""

import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración ──────────────────────────────────────────────────────────────
HTML_PATH   = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Descargas FBref/Temporada Pasada/laliga24-25.html"
OUTPUT_DIR  = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Temporada Pasada"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Datos FBref 24-25.xlsx")

# ── Leer HTML ──────────────────────────────────────────────────────────────────
with open(HTML_PATH, "r", encoding="utf-8") as f:
    content = f.read()

# ── Extraer tablas ─────────────────────────────────────────────────────────────
def get_table(table_id):
    return pd.read_html(io.StringIO(content), attrs={"id": table_id})[0]

# NOTA: Los IDs de las tablas de resultados incluyen el año de la temporada.
# Si los IDs siguientes no funcionan, ábrela el HTML en el navegador e inspecciona
# el atributo id de las tablas de clasificación para ajustarlos.
df_overall   = get_table("results2024-2025121_overall")
df_home_away = get_table("results2024-2025121_home_away")
df_std_for   = get_table("stats_squads_standard_for")
df_std_opp   = get_table("stats_squads_standard_against")
df_gk        = get_table("stats_squads_keeper_for")
df_misc_for  = get_table("stats_squads_misc_for")
df_misc_opp  = get_table("stats_squads_misc_against")

# ── Aplanar multi-index de columnas ───────────────────────────────────────────
def flatten_columns(df):
    new_cols = []
    for col in df.columns:
        if isinstance(col, tuple):
            top, sub = col
            new_cols.append(sub if top.startswith("Unnamed") else f"{top}_{sub}")
        else:
            new_cols.append(col)
    df.columns = new_cols
    return df

df_home_away = flatten_columns(df_home_away)
df_std_for   = flatten_columns(df_std_for)
df_std_opp   = flatten_columns(df_std_opp)
df_gk        = flatten_columns(df_gk)
df_misc_for  = flatten_columns(df_misc_for)
df_misc_opp  = flatten_columns(df_misc_opp)

# ── MP por equipo desde Tabla General ─────────────────────────────────────────
df_overall = df_overall.drop(columns=["Attendance", "Notes"], errors="ignore")
mp_map = df_overall.set_index("Squad")["MP"]

# ── Tabla L-V ─────────────────────────────────────────────────────────────────
df_home_away = df_home_away.drop(columns=["Rk"], errors="ignore")

# ── Squad Standard (for & opp) ────────────────────────────────────────────────
COLS_TO_DROP_STD = [
    "Playing Time_MP", "Playing Time_Starts", "Playing Time_Min", "Playing Time_90s",
    "Performance_G+A", "Per 90 Minutes_G+A", "Per 90 Minutes_G+A-PK",
]
df_std_for = df_std_for.drop(columns=COLS_TO_DROP_STD, errors="ignore")
df_std_opp = df_std_opp.drop(columns=COLS_TO_DROP_STD, errors="ignore")

def enrich_standard(df_std, df_misc, mp_map, is_opp=False):
    if is_opp:
        df_std  = df_std.copy()
        df_misc = df_misc.copy()
        df_std["_squad_key"]  = df_std["Squad"].str.replace("^vs ", "", regex=True)
        df_misc["_squad_key"] = df_misc["Squad"].str.replace("^vs ", "", regex=True)
        merge_key = "_squad_key"
    else:
        merge_key = "Squad"

    misc_cols = df_misc[[merge_key, "Performance_2CrdY", "Performance_Off"]].copy()
    df_std = df_std.merge(misc_cols, on=merge_key, how="left")
    df_std["MP"] = df_std[merge_key].map(mp_map)

    df_std["Performance_PK%"]      = (df_std["Performance_PK"] / df_std["Performance_PKatt"]).round(3)
    df_std["Per 90 Minutes_PK"]    = (df_std["Performance_PK"]    / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_PKatt"] = (df_std["Performance_PKatt"] / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_CrdY"]  = (df_std["Performance_CrdY"]  / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_CrdR"]  = (df_std["Performance_CrdR"]  / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_Off"]   = (df_std["Performance_Off"]   / df_std["MP"]).round(3)

    df_std = df_std.drop(columns=["_squad_key", "MP"], errors="ignore")

    if is_opp:
        ordered_cols = [
            "Squad",
            "Performance_Gls", "Performance_Ast",
            "Performance_G-PK", "Performance_PK", "Performance_PKatt", "Performance_PK%",
            "Performance_CrdY", "Performance_CrdR", "Performance_2CrdY", "Performance_Off",
            "Per 90 Minutes_Gls", "Per 90 Minutes_Ast", "Per 90 Minutes_G-PK",
            "Per 90 Minutes_PK", "Per 90 Minutes_PKatt",
            "Per 90 Minutes_CrdY", "Per 90 Minutes_CrdR", "Per 90 Minutes_Off",
        ]
    else:
        ordered_cols = [
            "Squad", "# Pl", "Age", "Poss",
            "Performance_Gls", "Performance_Ast",
            "Performance_G-PK", "Performance_PK", "Performance_PKatt", "Performance_PK%",
            "Performance_CrdY", "Performance_CrdR", "Performance_2CrdY", "Performance_Off",
            "Per 90 Minutes_Gls", "Per 90 Minutes_Ast", "Per 90 Minutes_G-PK",
            "Per 90 Minutes_PK", "Per 90 Minutes_PKatt",
            "Per 90 Minutes_CrdY", "Per 90 Minutes_CrdR", "Per 90 Minutes_Off",
        ]
    return df_std[[c for c in ordered_cols if c in df_std.columns]]

df_std_for = enrich_standard(df_std_for, df_misc_for, mp_map, is_opp=False)
df_std_opp = enrich_standard(df_std_opp, df_misc_opp, mp_map, is_opp=True)

# ── Goalkeeping ───────────────────────────────────────────────────────────────
df_gk = df_gk.drop(columns=[
    "Playing Time_MP", "Playing Time_Starts", "Playing Time_Min", "Playing Time_90s",
    "Performance_GA", "Performance_GA90", "Performance_W", "Performance_D", "Performance_L",
], errors="ignore")

df_gk["MP"] = df_gk["Squad"].map(mp_map)
df_gk["Performance_SoTA/MP"]  = (df_gk["Performance_SoTA"]  / df_gk["MP"]).round(3)
df_gk["Performance_Saves/MP"] = (df_gk["Performance_Saves"] / df_gk["MP"]).round(3)
df_gk = df_gk.drop(columns=["MP"], errors="ignore")

gk_ordered_cols = [
    "Squad", "# Pl",
    "Performance_SoTA", "Performance_SoTA/MP",
    "Performance_Saves", "Performance_Saves/MP",
    "Performance_Save%",
    "Performance_CS", "Performance_CS%",
    "Penalty Kicks_PKatt", "Penalty Kicks_PKA", "Penalty Kicks_PKsv",
    "Penalty Kicks_PKm", "Penalty Kicks_Save%",
]
df_gk = df_gk[[c for c in gk_ordered_cols if c in df_gk.columns]]

# ── Helpers de estilo ──────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
thin         = Side(style="thin", color="AAAAAA")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_df_to_sheet(ws, df, title=None):
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(name="Arial", bold=True, size=12)
        start_row = 2

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        fill = ALT_FILL if (row_idx - start_row) % 2 == 0 else WHITE_FILL
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.border    = BORDER
            if isinstance(value, str):
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].fillna("").astype(str).str.len().max() if not df.empty else 0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ── Crear workbook ─────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

ws1 = wb.create_sheet("Tabla General")
write_df_to_sheet(ws1, df_overall, title="La Liga 2024-2025 — Tabla General")

ws2 = wb.create_sheet("Tabla L-V")
write_df_to_sheet(ws2, df_home_away, title="La Liga 2024-2025 — Tabla Local / Visitante")

ws3 = wb.create_sheet("Squad Standard")
write_df_to_sheet(ws3, df_std_for, title="La Liga 2024-2025 — Squad Standard Stats (For)")

ws4 = wb.create_sheet("Squad Standard Opp")
write_df_to_sheet(ws4, df_std_opp, title="La Liga 2024-2025 — Squad Standard Stats (Against)")

ws5 = wb.create_sheet("Goalkeeping")
write_df_to_sheet(ws5, df_gk, title="La Liga 2024-2025 — Squad Goalkeeping")

# ── Guardar ────────────────────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"✅ Excel guardado en: {OUTPUT_FILE}")
