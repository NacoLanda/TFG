"""
Código Partidos 24-25.py
=========================
Versión de la temporada pasada (LaLiga 2024-25) del script Código Partidos.py.
Genera Partidos.xlsx en la carpeta Datos/Temporada Pasada/.

Ver Código Partidos.py (temporada actual) para documentación completa.

Uso: python3 "Temporada Pasada/Código Partidos 24-25.py"
"""

import pandas as pd
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ──────────────────────────────────────────────────────────────────────
HTML_PATH    = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Descargas FBref/Temporada Pasada/Partidos24-25.html"
OUTPUT_DIR   = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Temporada Pasada"
OUTPUT_FILE  = os.path.join(OUTPUT_DIR, "Partidos 24-25.xlsx")

# Columnas que NO queremos en el Excel final
COLS_EXCLUIR = {"Attendance", "Venue", "Match Report", "Notes"}

# ── Leer el HTML ───────────────────────────────────────────────────────────────
with open(HTML_PATH, "r", encoding="utf-8") as f:
    soup = BeautifulSoup(f, "html.parser")

table = soup.find_all("table")[9]
headers = [th.get_text(strip=True) for th in table.find("thead").find_all("th")]

rows = []
for tr in table.find("tbody").find_all("tr"):
    cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
    if cells:
        rows.append(cells)

df = pd.DataFrame(rows, columns=headers)

# ── Eliminar columnas no deseadas ──────────────────────────────────────────────
df.drop(columns=[c for c in df.columns if c in COLS_EXCLUIR], inplace=True)

# ── Eliminar filas donde Wk no sea un número ───────────────────────────────────
df = df[pd.to_numeric(df["Wk"], errors="coerce").notna()]

# ── Eliminar filas donde Score esté vacío ─────────────────────────────────────
df = df[df["Score"].str.strip() != ""]

df.replace("", pd.NA, inplace=True)
df.dropna(how="all", inplace=True)
df.fillna("", inplace=True)

# ── Crear directorio de salida si no existe ────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Escribir el Excel con formato profesional ──────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Partidos"

COLOR_HDR_BG  = "1F4E79"
COLOR_HDR_FT  = "FFFFFF"
COLOR_ROW_ALT = "D6E4F0"
BORDER_COLOR  = "BFBFBF"

thin   = Side(style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Cabeceras
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font      = Font(name="Arial", bold=True, color=COLOR_HDR_FT, size=11)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border
ws.row_dimensions[1].height = 30

# Datos
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    bg   = COLOR_ROW_ALT if row_idx % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=bg)
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

# Ajustar ancho de columnas automáticamente
for col_idx, col_name in enumerate(df.columns, start=1):
    max_len = len(str(col_name))
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

# Inmovilizar la fila de cabecera
ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"✅ Excel guardado en: {OUTPUT_FILE}")
print(f"   Filas exportadas : {len(df)}")
print(f"   Columnas         : {list(df.columns)}")
