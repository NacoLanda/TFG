"""
Código FBref.py
===============
Parsea el HTML descargado de FBref (LaLiga 2025-26) y genera un Excel
con cinco pestañas de estadísticas de equipo.

Flujo:
  1. Lee el archivo laliga.html guardado localmente desde FBref.
  2. Extrae 7 tablas identificadas por su id HTML.
  3. Aplana los multi-índices de columnas (FBref usa cabeceras en dos filas).
  4. Calcula métricas derivadas por partido (PK%, tarjetas/MP, fueras de juego/MP).
  5. Escribe un Excel formateado en Datos/Datos FBref.xlsx.

Entrada:  Descargas FBref/laliga.html  (descargado manualmente desde fbref.com)
Salida:   Datos/Datos FBref.xlsx       (5 hojas: Tabla General, Tabla L-V,
           Squad Standard, Squad Standard Opp, Goalkeeping)

Uso: python3 "Código FBref.py"
"""

import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — rutas de entrada y salida
# ══════════════════════════════════════════════════════════════════
HTML_PATH   = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Descargas FBref/laliga.html"
OUTPUT_DIR  = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Datos FBref.xlsx")

# ══════════════════════════════════════════════════════════════════
# LECTURA DEL HTML
# Cargamos todo el contenido en memoria para reutilizarlo en
# múltiples llamadas a pd.read_html sin releer el archivo.
# ══════════════════════════════════════════════════════════════════
with open(HTML_PATH, "r", encoding="utf-8") as f:
    content = f.read()

# ══════════════════════════════════════════════════════════════════
# EXTRACCIÓN DE TABLAS
# Cada tabla se identifica por su id HTML tal como aparece en FBref.
# ══════════════════════════════════════════════════════════════════
def get_table(table_id):
    """Extrae una tabla del HTML usando su id de elemento."""
    return pd.read_html(io.StringIO(content), attrs={"id": table_id})[0]

df_overall   = get_table("results2025-2026121_overall")    # clasificación general
df_home_away = get_table("results2025-2026121_home_away")  # resultados local/visitante
df_std_for   = get_table("stats_squads_standard_for")      # estadísticas estándar (propio)
df_std_opp   = get_table("stats_squads_standard_against")  # estadísticas estándar (rival)
df_gk        = get_table("stats_squads_keeper_for")        # portería
df_misc_for  = get_table("stats_squads_misc_for")          # miscelánea (propio)
df_misc_opp  = get_table("stats_squads_misc_against")      # miscelánea (rival)

# ══════════════════════════════════════════════════════════════════
# APLANADO DE MULTI-ÍNDICE
# FBref usa columnas en dos niveles (ej. "Performance" / "Gls").
# Las convertimos a una sola cadena: "Performance_Gls".
# Las columnas sin grupo se dejan con solo el nombre del sub-nivel.
# ══════════════════════════════════════════════════════════════════
def flatten_columns(df):
    """Convierte el multi-índice de columnas de FBref a nombres planos."""
    new_cols = []
    for col in df.columns:
        if isinstance(col, tuple):
            top, sub = col
            new_cols.append(sub if top.startswith("Unnamed") else f"{top}_{sub}")
        else:
            new_cols.append(col)
    df.columns = new_cols
    return df

df_home_away = flatten_columns(df_home_away)
df_std_for   = flatten_columns(df_std_for)
df_std_opp   = flatten_columns(df_std_opp)
df_gk        = flatten_columns(df_gk)
df_misc_for  = flatten_columns(df_misc_for)
df_misc_opp  = flatten_columns(df_misc_opp)

# ── Partidos jugados (MP) por equipo, necesario para calcular ratios /MP ───────
df_overall = df_overall.drop(columns=["Attendance", "Notes"], errors="ignore")
mp_map = df_overall.set_index("Squad")["MP"]  # dict equipo → partidos jugados

# ── Tabla Local-Visitante: eliminamos columna de ranking (innecesaria) ─────────
df_home_away = df_home_away.drop(columns=["Rk"], errors="ignore")

# ══════════════════════════════════════════════════════════════════
# ESTADÍSTICAS ESTÁNDAR (Squad Standard for / against)
# Eliminamos columnas redundantes con tiempo de juego y G+A
# (ya aparecen en otras tablas). Luego enriquecemos con métricas
# derivadas: ratios de tarjetas, fueras de juego y penaltis por MP.
# ══════════════════════════════════════════════════════════════════
COLS_TO_DROP_STD = [
    "Playing Time_MP", "Playing Time_Starts", "Playing Time_Min", "Playing Time_90s",
    "Performance_G+A", "Per 90 Minutes_G+A", "Per 90 Minutes_G+A-PK",
]
df_std_for = df_std_for.drop(columns=COLS_TO_DROP_STD, errors="ignore")
df_std_opp = df_std_opp.drop(columns=COLS_TO_DROP_STD, errors="ignore")

def enrich_standard(df_std, df_misc, mp_map, is_opp=False):
    """
    Añade métricas derivadas a la tabla estándar de equipo.

    Fusiona columnas disciplinarias de la tabla 'misc' y calcula
    ratios por partido (tarjetas, penaltis, fueras de juego).

    Args:
        df_std:  DataFrame de estadísticas estándar (Squad Standard).
        df_misc: DataFrame de miscelánea con tarjetas de 2ª amarilla y fueras de juego.
        mp_map:  Serie con el número de partidos jugados por equipo.
        is_opp:  True si es la tabla del rival (los nombres de equipo llevan prefijo "vs ").
    """
    if is_opp:
        df_std  = df_std.copy()
        df_misc = df_misc.copy()
        df_std["_squad_key"]  = df_std["Squad"].str.replace("^vs ", "", regex=True)
        df_misc["_squad_key"] = df_misc["Squad"].str.replace("^vs ", "", regex=True)
        merge_key = "_squad_key"
    else:
        merge_key = "Squad"

    misc_cols = df_misc[[merge_key, "Performance_2CrdY", "Performance_Off"]].copy()
    df_std = df_std.merge(misc_cols, on=merge_key, how="left")
    df_std["MP"] = df_std[merge_key].map(mp_map)

    df_std["Performance_PK%"]      = (df_std["Performance_PK"] / df_std["Performance_PKatt"]).round(3)
    df_std["Per 90 Minutes_PK"]    = (df_std["Performance_PK"]    / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_PKatt"] = (df_std["Performance_PKatt"] / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_CrdY"]  = (df_std["Performance_CrdY"]  / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_CrdR"]  = (df_std["Performance_CrdR"]  / df_std["MP"]).round(3)
    df_std["Per 90 Minutes_Off"]   = (df_std["Performance_Off"]   / df_std["MP"]).round(3)

    df_std = df_std.drop(columns=["_squad_key", "MP"], errors="ignore")

    if is_opp:
        ordered_cols = [
            "Squad",
            "Performance_Gls", "Performance_Ast",
            "Performance_G-PK", "Performance_PK", "Performance_PKatt", "Performance_PK%",
            "Performance_CrdY", "Performance_CrdR", "Performance_2CrdY", "Performance_Off",
            "Per 90 Minutes_Gls", "Per 90 Minutes_Ast", "Per 90 Minutes_G-PK",
            "Per 90 Minutes_PK", "Per 90 Minutes_PKatt",
            "Per 90 Minutes_CrdY", "Per 90 Minutes_CrdR", "Per 90 Minutes_Off",
        ]
    else:
        ordered_cols = [
            "Squad", "# Pl", "Age", "Poss",
            "Performance_Gls", "Performance_Ast",
            "Performance_G-PK", "Performance_PK", "Performance_PKatt", "Performance_PK%",
            "Performance_CrdY", "Performance_CrdR", "Performance_2CrdY", "Performance_Off",
            "Per 90 Minutes_Gls", "Per 90 Minutes_Ast", "Per 90 Minutes_G-PK",
            "Per 90 Minutes_PK", "Per 90 Minutes_PKatt",
            "Per 90 Minutes_CrdY", "Per 90 Minutes_CrdR", "Per 90 Minutes_Off",
        ]
    return df_std[[c for c in ordered_cols if c in df_std.columns]]

df_std_for = enrich_standard(df_std_for, df_misc_for, mp_map, is_opp=False)
df_std_opp = enrich_standard(df_std_opp, df_misc_opp, mp_map, is_opp=True)

# ══════════════════════════════════════════════════════════════════
# ESTADÍSTICAS DE PORTERÍA (Goalkeeping)
# Eliminamos columnas redundantes y añadimos ratios de tiros
# y paradas por partido (SoTA/MP y Saves/MP).
# ══════════════════════════════════════════════════════════════════
df_gk = df_gk.drop(columns=[
    "Playing Time_MP", "Playing Time_Starts", "Playing Time_Min", "Playing Time_90s",
    "Performance_GA", "Performance_GA90", "Performance_W", "Performance_D", "Performance_L",
], errors="ignore")

df_gk["MP"] = df_gk["Squad"].map(mp_map)
df_gk["Performance_SoTA/MP"]  = (df_gk["Performance_SoTA"]  / df_gk["MP"]).round(3)
df_gk["Performance_Saves/MP"] = (df_gk["Performance_Saves"] / df_gk["MP"]).round(3)
df_gk = df_gk.drop(columns=["MP"], errors="ignore")

gk_ordered_cols = [
    "Squad", "# Pl",
    "Performance_SoTA", "Performance_SoTA/MP",
    "Performance_Saves", "Performance_Saves/MP",
    "Performance_Save%",
    "Performance_CS", "Performance_CS%",
    "Penalty Kicks_PKatt", "Penalty Kicks_PKA", "Penalty Kicks_PKsv",
    "Penalty Kicks_PKm", "Penalty Kicks_Save%",
]
df_gk = df_gk[[c for c in gk_ordered_cols if c in df_gk.columns]]

# ══════════════════════════════════════════════════════════════════
# ESTILOS EXCEL — constantes reutilizadas en todas las hojas
# ══════════════════════════════════════════════════════════════════
HEADER_FILL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
thin         = Side(style="thin", color="AAAAAA")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_df_to_sheet(ws, df, title=None):
    """
    Escribe un DataFrame en una hoja de Excel con formato profesional.

    Aplica cabecera con fondo azul oscuro, filas alternadas y bordes.
    Ajusta el ancho de cada columna al contenido y fija la primera fila.

    Args:
        ws:    Hoja de Excel (openpyxl Worksheet).
        df:    DataFrame a escribir.
        title: Si se indica, se escribe en la celda A1 como título de sección.
    """
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(name="Arial", bold=True, size=12)
        start_row = 2

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        fill = ALT_FILL if (row_idx - start_row) % 2 == 0 else WHITE_FILL
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.border    = BORDER
            if isinstance(value, str):
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].fillna("").astype(str).str.len().max() if not df.empty else 0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ══════════════════════════════════════════════════════════════════
# ESCRITURA DEL EXCEL
# Creamos el workbook y volcamos cada tabla en su propia hoja.
# ══════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

ws1 = wb.create_sheet("Tabla General")
write_df_to_sheet(ws1, df_overall, title="La Liga 2025-2026 — Tabla General")

ws2 = wb.create_sheet("Tabla L-V")
write_df_to_sheet(ws2, df_home_away, title="La Liga 2025-2026 — Tabla Local / Visitante")

ws3 = wb.create_sheet("Squad Standard")
write_df_to_sheet(ws3, df_std_for, title="La Liga 2025-2026 — Squad Standard Stats (For)")

ws4 = wb.create_sheet("Squad Standard Opp")
write_df_to_sheet(ws4, df_std_opp, title="La Liga 2025-2026 — Squad Standard Stats (Against)")

ws5 = wb.create_sheet("Goalkeeping")
write_df_to_sheet(ws5, df_gk, title="La Liga 2025-2026 — Squad Goalkeeping")

# ── Guardar en disco ───────────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"✅ Excel guardado en: {OUTPUT_FILE}")