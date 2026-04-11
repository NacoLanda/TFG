"""
Código Partidos.py
==================
Parsea el HTML de resultados de FBref y lo enriquece con datos de lluvia
por partido, generando un Excel con el calendario completo de LaLiga 2025-26.

Flujo:
  1. Lee Partidos.html (descargado manualmente de FBref) y extrae la tabla
     de resultados (tabla índice 9 del HTML).
  2. Limpia filas vacías y sin marcador.
  3. Para cada partido, busca en Datos Lluvias.xlsx si llovió durante las
     horas del partido (cruce por fecha + equipo local + horario).
  4. Añade la columna "Lluvia" (mm registrados o "No Llovió").
  5. Escribe un Excel formateado en Datos/Partidos.xlsx.

NOTA: Este script se ejecuta DOS veces en Actualización.py de forma intencionada:
  - Primera pasada: genera Partidos.xlsx sin lluvia (lo usa Código Lluvias.py).
  - Segunda pasada: añade los datos de lluvia ya disponibles.

Entrada:  Descargas FBref/Partidos.html  y  Datos/Datos Lluvias.xlsx
Salida:   Datos/Partidos.xlsx

Uso: python3 "Código Partidos.py"
"""

import pandas as pd
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — rutas y columnas a excluir
# ══════════════════════════════════════════════════════════════════
HTML_PATH    = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Descargas FBref/Partidos.html"
LLUVIA_PATH  = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Datos Lluvias.xlsx"
OUTPUT_DIR   = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos"
OUTPUT_FILE  = os.path.join(OUTPUT_DIR, "Partidos.xlsx")

# Columnas que NO queremos en el Excel final
COLS_EXCLUIR = {"Attendance", "Venue", "Match Report", "Notes"}

# ══════════════════════════════════════════════════════════════════
# EXTRACCIÓN DE PARTIDOS DESDE HTML
# FBref incluye varias tablas en la página; la de resultados es la
# número 9 (índice base 0). Se extraen cabeceras y filas manualmente
# porque pd.read_html no maneja bien las filas de separación de jornadas.
# ══════════════════════════════════════════════════════════════════
with open(HTML_PATH, "r", encoding="utf-8") as f:
    soup = BeautifulSoup(f, "html.parser")

# La tabla de resultados es la décima tabla del HTML (índice 9).
# Las otras tablas de la misma página contienen información de navegación
# y clasificaciones que no interesan.
table = soup.find_all("table")[9]
# La cabecera usa <th> dentro de <thead>; los datos usan <td> dentro de <tbody>
headers = [th.get_text(strip=True) for th in table.find("thead").find_all("th")]

rows = []
for tr in table.find("tbody").find_all("tr"):
    # Algunas filas solo tienen <th> (separadores de jornada); se recogen igual
    # y se filtrarán después por si la columna "Wk" no es numérica.
    cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
    if cells:
        rows.append(cells)

df = pd.DataFrame(rows, columns=headers)

# ── Limpieza: eliminar columnas no deseadas, filas de cabecera y sin resultado ──
df.drop(columns=[c for c in df.columns if c in COLS_EXCLUIR], inplace=True)

# FBref intercala filas de cabecera dentro del tbody; las filtramos
# comprobando que la columna "Wk" (jornada) contenga un número.
df = df[pd.to_numeric(df["Wk"], errors="coerce").notna()]

# Partidos no jugados aún tienen el Score vacío; los descartamos.
df = df[df["Score"].str.strip() != ""]

df.replace("", pd.NA, inplace=True)
df.dropna(how="all", inplace=True)
df.fillna("", inplace=True)

# ══════════════════════════════════════════════════════════════════
# ENRIQUECIMIENTO CON DATOS DE LLUVIA
# Cruzamos cada partido (fecha + equipo local + horario) con el
# registro horario de precipitaciones de Datos Lluvias.xlsx.
# Si no existe ese archivo todavía (primera pasada), la columna
# "Lluvia" quedará como "No Llovió" para todos los partidos.
# ══════════════════════════════════════════════════════════════════
df_lluvia = pd.read_excel(LLUVIA_PATH, sheet_name="Datos diarios")
df_lluvia["Fecha"] = pd.to_datetime(df_lluvia["Fecha"], dayfirst=True).dt.date
df_lluvia["_idx"]  = range(len(df_lluvia))

# Índice (fecha, club_local) → índice de fila en df_lluvia
lluvia_index = {}
for _, row in df_lluvia.iterrows():
    key = (row["Fecha"], str(row["Club local"]).strip())
    lluvia_index[key] = row["_idx"]

def obtener_lluvia(row):
    """
    Determina si llovió durante un partido consultando los datos horarios.

    Busca el partido en el índice (fecha, club_local) y extrae los mm
    registrados en las horas del partido (inicio e inicio+1).

    Devuelve:
        "No Llovió"        — sin precipitación en esas horas
        "Estadio Cubierto" — estadio con techo
        "X,XX mm"          — milímetros medios registrados durante el partido
    """
    try:
        fecha         = pd.to_datetime(row["Date"]).date()
        club          = str(row["Home"]).strip()
        hora_inicio   = int(str(row["Time"]).strip().split(":")[0])
        horas_partido = {hora_inicio, hora_inicio + 1}

        idx = lluvia_index.get((fecha, club))
        if idx is None:
            return "No Llovió"

        valor = df_lluvia.loc[df_lluvia["_idx"] == idx, "Horas con lluvia (hora: mm)"].values[0]

        if pd.isna(valor) or str(valor).strip() == "Sin lluvia":
            return "No Llovió"
        if "Estadio Cubierto" in str(valor):
            return "Estadio Cubierto"

        # Parsear entradas tipo "19:00: 2.3 mm | 20:00: 1.1 mm"
        mm_partido = []
        for entrada in str(valor).split("|"):
            entrada = entrada.strip()
            try:
                partes = entrada.split(":")
                hora   = int(partes[0].strip())
                mm     = float(partes[-1].replace("mm", "").strip())
                if hora in horas_partido:
                    mm_partido.append(mm)
            except Exception:
                continue

        if not mm_partido:
            return "No Llovió"
        media = round(sum(mm_partido) / len(mm_partido), 2)
        return f"{str(media).replace('.', ',')} mm"

    except Exception:
        return "No Llovió"

df["Lluvia"] = df.apply(obtener_lluvia, axis=1)

# ══════════════════════════════════════════════════════════════════
# ESCRITURA DEL EXCEL
# ══════════════════════════════════════════════════════════════════
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "Partidos"

COLOR_HDR_BG  = "1F4E79"
COLOR_HDR_FT  = "FFFFFF"
COLOR_ROW_ALT = "D6E4F0"
BORDER_COLOR  = "BFBFBF"

thin   = Side(style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Cabeceras
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font      = Font(name="Arial", bold=True, color=COLOR_HDR_FT, size=11)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border
ws.row_dimensions[1].height = 30

# Datos
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    bg   = COLOR_ROW_ALT if row_idx % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=bg)
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

# Ajustar ancho de columnas automáticamente
for col_idx, col_name in enumerate(df.columns, start=1):
    max_len = len(str(col_name))
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

# Inmovilizar la fila de cabecera
ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"✅ Excel guardado en: {OUTPUT_FILE}")
print(f"   Filas exportadas : {len(df)}")
print(f"   Columnas         : {list(df.columns)}")
