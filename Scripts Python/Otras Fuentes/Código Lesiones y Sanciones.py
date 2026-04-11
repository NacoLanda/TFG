"""
Código Lesiones y Sanciones.py
================================
Descarga en tiempo real la lista de jugadores lesionados y sancionados
de LaLiga desde futbolfantasy.com y genera un Excel con dos hojas.

Flujo:
  1. Scraping de la página de lesionados: extrae jugador, equipo, tipo
     de lesión, fecha de inicio, días lesionado, vuelta esperada y
     probabilidad de jugar.
  2. Scraping de la página de sancionados: extrae jugador, equipo y motivo.
  3. Escribe ambas listas en Datos/Lesiones y Sanciones.xlsx.

Salida: Datos/Lesiones y Sanciones.xlsx  (2 hojas: Lesiones y Sanciones)

Uso: python3 "Código Lesiones y Sanciones.py"
"""

import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Cabecera HTTP para simular un navegador y evitar bloqueos del servidor
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

# Normalización de nombres de equipo: la web usa nombres cortos,
# los homogeneizamos al formato usado en el resto del proyecto.
equipos_map = {
    "Athletic": "Athletic Club",
    "Atlético": "Atlético Madrid",
    "Betis": "Real Betis",
    "Celta": "Celta Vigo",
    "Rayo": "Rayo Vallecano",
    "Real Oviedo": "Oviedo",
}

# ══════════════════════════════════════════════════════════════════
# SCRAPING DE LESIONADOS
# La página agrupa jugadores por equipo: primero aparece un <img>
# con el escudo del equipo (cuya URL contiene "cabecera/hd/"),
# y a continuación los <a> con los jugadores lesionados de ese equipo.
# Recorremos todos los tags en orden para mantener el contexto de equipo.
# ══════════════════════════════════════════════════════════════════
URL_LES = "https://www.futbolfantasy.com/laliga/lesionados"
response_les = requests.get(URL_LES, headers=headers)
soup_les = BeautifulSoup(response_les.text, "html.parser")

data_les = []
equipo_actual = ""

# Recorremos todos los tags img y a en orden de aparición para mantener
# el contexto del equipo actual: cuando encontramos un escudo (img con
# "cabecera/hd/" en la URL) actualizamos equipo_actual; cuando encontramos
# un enlace a un jugador (/jugadores/), extraemos sus datos con ese equipo.
for tag in soup_les.find_all(["img", "a"]):
    if tag.name == "img" and "cabecera/hd/" in (tag.get("src") or ""):
        # El escudo del equipo está dentro de un elemento con el nombre del equipo
        padre = tag.find_parent()
        if padre:
            equipo_actual = padre.get_text(strip=True)
        if not equipo_actual:
            equipo_actual = tag.get("alt", "").strip()
        equipo_actual = equipos_map.get(equipo_actual, equipo_actual)

    if tag.name == "a" and "/jugadores/" in (tag.get("href") or ""):
        nombre = tag.get_text(strip=True)
        if not nombre:
            continue

        parent = tag.find_parent()
        lesion = ""
        fecha_inicio = ""
        dias_lesionado = ""
        vuelta_esperada = ""
        probabilidad = ""

        if parent:
            texts = [t.strip() for t in parent.stripped_strings]

            # El tipo de lesión está en el span inmediatamente posterior al enlace
            siguiente_span = tag.find_next("span")
            if siguiente_span:
                lesion = siguiente_span.get_text(strip=True)

            # La fecha de inicio y los días lesionado están en el texto del bloque
            for t in texts:
                match_fecha = re.search(r"Desde\s+(\d{2}/\d{2})", t)
                if match_fecha:
                    fecha_inicio = match_fecha.group(1)

                match_dias = re.search(r"\((\d+)\s+días?\)", t)
                if match_dias:
                    dias_lesionado = match_dias.group(1) + " días"

            # La vuelta esperada está en el span que sigue al span con los días
            span_dias = None
            for span in parent.find_all("span"):
                if re.search(r"\d+\s+días?", span.get_text()):
                    span_dias = span
                    break
            if span_dias:
                siguiente_span_vuelta = span_dias.find_next("span")
                if siguiente_span_vuelta:
                    vuelta_esperada = siguiente_span_vuelta.get_text(strip=True)

            # La probabilidad de jugar está en el span de clase "probabilidad-widget"
            # que aparece ANTES del enlace del jugador (encabeza el bloque)
            span_prob = tag.find_previous("span", class_="probabilidad-widget")
            if span_prob:
                probabilidad = span_prob.get_text(strip=True)

        data_les.append({
            "Jugador": nombre,
            "Equipo": equipo_actual,
            "Lesión": lesion,
            "Fecha inicio": fecha_inicio,
            "Tiempo lesionado": dias_lesionado,
            "Vuelta esperada": vuelta_esperada,
            "Probabilidad de jugar": probabilidad,
        })

# ══════════════════════════════════════════════════════════════════
# SCRAPING DE SANCIONADOS
# Misma lógica que lesionados: escudo → equipo, enlace → jugador.
# Se extrae el motivo de la sanción desde el span con clase "sancion".
# ══════════════════════════════════════════════════════════════════
URL_SAN = "https://www.futbolfantasy.com/laliga/sancionados"
response_san = requests.get(URL_SAN, headers=headers)
soup_san = BeautifulSoup(response_san.text, "html.parser")

data_san = []
equipo_san = ""

# Misma lógica de recorrido que en lesionados: escudo → equipo, enlace → jugador.
for tag in soup_san.find_all(["img", "a"]):
    if tag.name == "img" and "cabecera/hd/" in (tag.get("src") or ""):
        padre = tag.find_parent()
        if padre:
            equipo_san = padre.get_text(strip=True)
        if not equipo_san:
            equipo_san = tag.get("alt", "").strip()
        equipo_san = equipos_map.get(equipo_san, equipo_san)

    if tag.name == "a" and "/jugadores/" in (tag.get("href") or ""):
        nombre = tag.get_text(strip=True)
        if not nombre:
            continue

        motivo = ""
        # El motivo de la sanción está en el span con clase "sancion" que sigue al enlace
        span_sancion = tag.find_next("span", class_="sancion")
        if span_sancion:
            # Se extrae solo el texto directo del span (sin descendientes)
            motivo = span_sancion.find(string=True, recursive=False)
            if motivo:
                motivo = motivo.strip()

        data_san.append({
            "Jugador": nombre,
            "Equipo": equipo_san,
            "Motivo": motivo,
        })

# ══════════════════════════════════════════════════════════════════
# ESCRITURA DEL EXCEL
# Dos hojas independientes: "Lesiones" y "Sanciones".
# ══════════════════════════════════════════════════════════════════
wb = Workbook()
COLOR_HEADER = "1F4E79"
thin = Side(style="thin", color="CCCCCC")

# ── Hoja 1: Lesionados ─────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Lesiones"

headers_les = ["Jugador", "Equipo", "Lesión", "Fecha inicio", "Tiempo lesionado", "Vuelta esperada", "Probabilidad de jugar"]
for col, h in enumerate(headers_les, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, item in enumerate(data_les, 2):
    for col, key in enumerate(headers_les, 1):
        cell = ws1.cell(row=row_idx, column=col, value=item[key])
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=7):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 38
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 30
ws1.column_dimensions["G"].width = 22
ws1.row_dimensions[1].height = 22

# ── Hoja 2: Sancionados ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Sanciones")

headers_san = ["Jugador", "Equipo", "Motivo"]
for col, h in enumerate(headers_san, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, item in enumerate(data_san, 2):
    for col, key in enumerate(headers_san, 1):
        cell = ws2.cell(row=row_idx, column=col, value=item[key])
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 30
ws2.row_dimensions[1].height = 22

wb.save("/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Lesiones y Sanciones.xlsx")
print(f"✅ Archivo 'Lesiones y Sanciones.xlsx' creado con {len(data_les)} lesionados y {len(data_san)} sancionados.")
