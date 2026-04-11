"""
Código Árbitros.py
==================
Extrae estadísticas de árbitros de LaLiga 2025-26 desde tres fuentes web
y las consolida en un Excel con dos hojas.

Fuentes scrapeadas:
  1. estadisticaslaliga.es → tabla general de árbitros + designaciones por partido
  2. transfermarkt.es      → año de debut en Primera División y segundas amarillas
  3. whoscored.com         → % victorias local, % derrotas local, % empates

Flujo:
  1. Scrapea estadisticaslaliga.es (requests + BeautifulSoup) para la tabla general.
  2. Scrapea estadisticaslaliga.es para las designaciones por jornada.
  3. Scrapea transfermarkt.es con Selenium headless (requiere JS).
  4. Scrapea whoscored.com con Chrome visible (detecta modo headless).
  5. Unifica las tres fuentes por nombre de árbitro (normalización de texto).
  6. Escribe el Excel con la hoja General (una fila por árbitro) y
     la hoja Partidos (una fila por partido, agrupada por árbitro).

Salida: Datos/Datos Árbitros.xlsx  (2 hojas: General y Partidos)

NOTA: WhoScored abre una ventana de Chrome visible brevemente para cargar
los datos.

Uso: python3 "Código Árbitros.py"
"""

import time
import os
import re
import unicodedata
import requests
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — rutas, constantes y mapas de normalización
# NOMBRE_MAP: corrige el nombre de árbitros que difieren entre fuentes.
# WS_SLUG_MAP: convierte el slug de la URL de WhoScored al nombre completo.
# COLUMN_LABELS: etiquetas legibles para las columnas del Excel final.
# ══════════════════════════════════════════════════════════════════
OUTPUT_DIR  = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Datos Árbitros.xlsx")

# Árbitro excluido porque sus datos son incompletos o inconsistentes entre fuentes
ARBITRO_EXCLUIDO = "victor garcia acosta"

os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
}

NOMBRE_MAP = {
    "ricardo de burgos bengoetxea": "De Burgos Bengoetxea",
}

SOURCE_COLORS = {
    "_esl": "2E75B6",
    "_tm":  "ED7D31",
    "_ws":  "70AD47",
}
ARBITRO_COLOR = "1F3864"

WS_SLUG_MAP = {
    "jesus-gil-manzano":                    "Jesús Gil Manzano",
    "iosu-galech-apezteguia":               "Iosu Galech Apezteguía",
    "miguel-angel-ortiz-arias":             "Miguel Ángel Ortiz Arias",
    "isidro-diaz-de-mera-escuderos":        "Isidro Díaz de Mera Escuderos",
    "javier-alberola-rojas":                "Javier Alberola Rojas",
    "miguel-sesma-espinosa":                "Miguel Sesma Espinosa",
    "ricardo-de-burgos-bengoetxea":         "De Burgos Bengoetxea",
    "cesar-soto-grado":                     "César Soto Grado",
    "alejandro-jose-hernandez-hernandez":   "Alejandro Hernández Hernández",
    "francisco-jose-hernandez-maeso":       "Francisco José Hernández Maeso",
    "victor-garcia-verdura":                "Víctor García Verdura",
    "adrian-cordero-vega":                  "Adrián Cordero Vega",
    "alejandro-quintero-gonzalez":          "Alejandro Quintero González",
    "alejandro-muniz-ruiz":                 "Alejandro Muñiz Ruiz",
    "jose-luis-munuera-montero":            "José Luis Munuera Montero",
    "jose-luis-guzman-mansilla":            "José Luis Guzmán Mansilla",
    "guillermo-cuadra-fernandez":           "Guillermo Cuadra Fernández",
    "mateo-busquets-ferrer":                "Mateo Busquets Ferrer",
    "juan-martinez-munuera":                "Juan Martínez Munuera",
    "jose-maria-sanchez-martinez":          "José María Sánchez Martínez",
}

COLUMNAS_TEXTO = {"Árbitro", "Debut_tm"}

# Nombres legibles para las columnas de la hoja General
# Clave: nombre interno (con sufijo)  →  Valor: etiqueta mostrada en Excel
COLUMN_LABELS = {
    # estadisticaslaliga.es
    "Partidos_esl":   "Partidos",
    "Faltas_esl":     "Faltas",
    "Faltas/P_esl":   "Faltas / Partido",
    "Penaltis_esl":   "Penaltis",
    "Pen/P_esl":      "Penaltis / Partido",
    "Amarillas_esl":  "Amarillas",
    "Ama/P_esl":      "Amarillas / Partido",
    "Rojas_esl":      "Rojas",
    "Rojas/P_esl":    "Rojas / Partido",
    # transfermarkt.es
    "Debut_tm":             "Debut",
    "2ª Amarillas_tm":      "2ª Amarillas",
    "ø 2ª Amarillas_tm":    "Prom. 2ª Amarillas",
    # whoscored.com
    "Ganados en casa %_ws":  "% Ganados local",
    "Perdidos en casa %_ws": "% Perdidos local",
    "Empates %_ws":          "% Empates",
}


# ══════════════════════════════════════════════════════════════════
# HELPERS DE NORMALIZACIÓN Y SCRAPING
# normalizar(): elimina tildes y normaliza espacios para comparar nombres
#               de árbitros que las distintas fuentes escriben diferente.
# build_driver(): crea Chrome con opciones anti-detección.
# cerrar_banners(): elimina overlays y acepta cookies para liberar la página.
# ══════════════════════════════════════════════════════════════════
def normalizar(texto):
    """Convierte un texto a minúsculas, elimina tildes y normaliza espacios."""
    texto = texto.lower().strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def extraer_anio(fecha):
    """Extrae el año de cuatro dígitos de una cadena de texto de fecha."""
    m = re.search(r"\b(\d{4})\b", str(fecha))
    return m.group(1) if m else fecha


def nombre_desde_href(href):
    """Extrae el nombre de un árbitro a partir del slug de su URL en WhoScored."""
    slug = href.rstrip("/").split("/")[-1]
    return WS_SLUG_MAP.get(slug, slug.replace("-", " ").title())


def to_number(valor):
    """Convierte un valor a float aceptando comas como separador decimal."""
    if isinstance(valor, (int, float)):
        return valor
    try:
        return float(str(valor).replace(",", "."))
    except (ValueError, TypeError):
        return valor


def build_driver(headless=True):
    """Crea un WebDriver de Chrome con opciones anti-detección configuradas."""
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=opts)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def cerrar_banners(driver):
    """Elimina overlays y acepta banners de cookies en la página activa."""
    driver.execute_script("""
        var overlays = document.querySelectorAll(
            '[style*="z-index: 2147483647"], [style*="z-index:2147483647"], .a__sc-np32r2-0'
        );
        overlays.forEach(function(el) { el.remove(); });
    """)
    time.sleep(0.5)
    for sel in ["button[title='Aceptar todo']", "#onetrust-accept-btn-handler"]:
        try:
            btns = driver.find_elements(By.CSS_SELECTOR, sel)
            for btn in btns:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.5)
                    break
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════
# FUENTE 1A — estadisticaslaliga.es (tabla general de árbitros)
# Extrae la tabla principal con partidos, faltas, amarillas, rojas
# y penaltis por árbitro. Usa requests (no requiere JS).
# ══════════════════════════════════════════════════════════════════
def scrape_estadisticaslaliga():
    """
    Extrae la tabla general de árbitros de estadisticaslaliga.es.

    Devuelve un DataFrame con Árbitro + métricas con sufijo _esl
    (Partidos_esl, Faltas_esl, Amarillas_esl, etc.).
    """
    url = "https://www.estadisticaslaliga.es/arbitros.php"
    print("  -> Scrapeando estadisticaslaliga.es (tabla arbitros) ...")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        target_table = soup.find("table", class_="tabla-ranking")

        if not target_table:
            print("    x No se encontro la tabla en estadisticaslaliga.es")
            return pd.DataFrame()

        thead = target_table.find("thead")
        col_names = [th.get_text(strip=True) for th in thead.find_all("th")]

        data = []
        for row in target_table.find("tbody").find_all("tr"):
            cells = row.find_all(["td", "th"])
            if not cells:
                continue
            link = cells[0].find("a", class_="link-arbitro")
            nombre = link.get_text(strip=True) if link else cells[0].get_text(strip=True)
            row_data = [nombre] + [c.get_text(strip=True) for c in cells[1:]]
            data.append(row_data)

        if not data:
            return pd.DataFrame()

        df = pd.DataFrame(data, columns=col_names[:len(data[0])])
        df = df.dropna(how="all").reset_index(drop=True)

        nombre_col = df.columns[0]
        df = df.rename(columns={nombre_col: "Árbitro"})
        df = df.rename(columns={c: f"{c}_esl" for c in df.columns if c != "Árbitro"})

        print(f"    + {len(df)} arbitros obtenidos de estadisticaslaliga.es")
        return df

    except Exception as e:
        print(f"    x Error en estadisticaslaliga.es: {e}")
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════
# FUENTE 1B — estadisticaslaliga.es (designaciones arbitrales)
# Extrae la tabla de designaciones por jornada con el árbitro asignado
# y las estadísticas (faltas, tarjetas, ratio F/T) para local y visitante.
# La estructura HTML usa <h3>Jornada X</h3> seguido de divs de partido.
# ══════════════════════════════════════════════════════════════════
def scrape_designaciones():
    """
    Scrapes las designaciones arbitrales de estadisticaslaliga.es.
    La nueva estructura usa <h3>Jornada X</h3> para marcar cada jornada,
    <p> con <a href="recibe_arbitros.php?pedido=..."> para el arbitro de cada
    partido, y una <table> con dos filas (local y visitante) a continuacion.
    Devuelve un DataFrame con columnas:
      Arbitro | Jornada | Equipo Local | Equipo Visitante |
      Faltas Local | Faltas Visitante |
      Tarjetas Local | Tarjetas Visitante |
      F/T Local | F/T Visitante
    """
    url = "https://www.estadisticaslaliga.es/arbitros.php"
    print("  -> Scrapeando estadisticaslaliga.es (designaciones arbitrales) ...")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        data = []

        def cell(cells, idx):
            return re.sub(r"\s+", " ", cells[idx].get_text(strip=True)).strip() if len(cells) > idx else ""

        # Estructura real:
        # <h3>Jornada X</h3>
        # <div>  ← contenedor de todos los partidos de esa jornada
        #   <div>  ← un partido
        #     <div>Árbitro del Encuentro⚖️ NOMBRE</div>
        #     <div><table>...</table></div>
        #   </div>
        #   ...
        # </div>
        for h3 in soup.find_all("h3", string=re.compile(r"Jornada")):
            m = re.search(r"Jornada\s+(\d+)", h3.get_text())
            if not m:
                continue
            jornada_actual = int(m.group(1))

            contenedor = h3.find_next_sibling("div")
            if not contenedor:
                continue

            for partido_div in contenedor.find_all("div", recursive=False):
                hijos = [h for h in partido_div.find_all("div", recursive=False)]
                if len(hijos) < 2:
                    continue

                # Primer div hijo: nombre del árbitro
                texto_arb = hijos[0].get_text(strip=True)
                nombre = re.sub(r"Árbitro del Encuentro", "", texto_arb)
                nombre = re.sub(r"[^\w\s\u00C0-\u017E]", "", nombre).strip()

                if not nombre or normalizar(nombre) == ARBITRO_EXCLUIDO:
                    continue

                # Segundo div hijo: tabla con equipos
                tabla = hijos[1].find("table")
                if not tabla:
                    continue

                data_rows = [r for r in tabla.find_all("tr") if r.find("td")]
                if len(data_rows) < 2:
                    continue

                local = data_rows[0].find_all("td")
                away  = data_rows[1].find_all("td")

                data.append({
                    "Árbitro":            nombre,
                    "Jornada":            jornada_actual,
                    "Equipo Local":       cell(local, 0),
                    "Equipo Visitante":   cell(away,  0),
                    "Faltas Local":       cell(local, 1),
                    "Faltas Visitante":   cell(away,  1),
                    "Tarjetas Local":     cell(local, 2),
                    "Tarjetas Visitante": cell(away,  2),
                    "F/T Local":          cell(local, 3),
                    "F/T Visitante":      cell(away,  3),
                })

        if not data:
            print("    x No se extrajeron datos de designaciones")
            return pd.DataFrame()

        df = pd.DataFrame(data)
        df = df.sort_values(["Jornada", "Árbitro"]).reset_index(drop=True)
        print(f"    + {len(df)} partidos obtenidos de designaciones arbitrales")
        return df

    except Exception as e:
        print(f"    x Error en designaciones: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════
# FUENTE 2 — transfermarkt.es
# Extrae año de debut en Primera División y segundas amarillas por árbitro.
# Requiere Selenium headless porque la tabla carga con JavaScript.
# ══════════════════════════════════════════════════════════════════
def scrape_transfermarkt():
    """
    Extrae datos de árbitros de Transfermarkt (debut en Primera División,
    segundas amarillas acumuladas y promedio de segundas amarillas por partido).

    Requiere Selenium porque la página carga el contenido con JavaScript.
    Devuelve un DataFrame con columnas: Árbitro, Debut_tm, 2ª Amarillas_tm,
    ø 2ª Amarillas_tm.
    """
    url = (
        "https://www.transfermarkt.es/laliga/schiedsrichter/"
        "wettbewerb/ES1/saison_id/2025/plus/1"
    )
    print("  -> Scrapeando transfermarkt.es (Selenium headless) ...")
    driver = None
    try:
        driver = build_driver(headless=True)
        driver.execute_cdp_cmd(
            "Network.setExtraHTTPHeaders",
            {"headers": {"Accept-Language": "es-ES,es;q=0.9"}},
        )
        driver.get(url)

        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "items")))
        time.sleep(2)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        table = soup.find("table", {"class": "items"})
        if not table:
            return pd.DataFrame()
        tbody = table.find("tbody")
        if not tbody:
            return pd.DataFrame()

        data = []
        for row in tbody.find_all("tr", recursive=False):
            cells = row.find_all("td", recursive=False)
            if not cells:
                continue
            nombre = ""
            hauptlink = cells[0].find("td", {"class": "hauptlink"})
            if hauptlink:
                a = hauptlink.find("a")
                if a:
                    nombre = a.get_text(strip=True)
            if not nombre:
                continue
            debut = ""
            if len(cells) > 1:
                a = cells[1].find("a")
                fecha = a.get_text(strip=True) if a else cells[1].get_text(strip=True)
                debut = extraer_anio(fecha)
            seg_amarilla   = cells[6].get_text(strip=True) if len(cells) > 6 else ""
            o_seg_amarilla = cells[7].get_text(strip=True) if len(cells) > 7 else ""
            data.append([nombre, debut, seg_amarilla, o_seg_amarilla])

        if not data:
            return pd.DataFrame()

        df = pd.DataFrame(data, columns=[
            "Árbitro", "Debut_tm", "2ª Amarillas_tm", "ø 2ª Amarillas_tm"
        ])
        print(f"    + {len(df)} arbitros obtenidos de transfermarkt.es")
        return df

    except Exception as e:
        print(f"    x Error en transfermarkt.es: {e}")
        return pd.DataFrame()
    finally:
        if driver:
            driver.quit()


# ══════════════════════════════════════════════════════════════════
# FUENTE 3 — whoscored.com (sección árbitros)
# Extrae % victorias local, % derrotas local y % empates por árbitro.
# Usa Chrome en modo visible porque WhoScored bloquea el modo headless.
# ══════════════════════════════════════════════════════════════════
def scrape_whoscored_resultados():
    """
    Extrae estadísticas de árbitros de WhoScored (sección de árbitros).

    Abre Chrome visible (no headless) porque WhoScored detecta y bloquea
    el modo headless. Extrae % victorias local, % derrotas local y % empates
    para cada árbitro de la temporada.

    Devuelve un DataFrame con columnas: Árbitro, Ganados en casa %_ws,
    Perdidos en casa %_ws, Empates %_ws.
    """
    url = (
        "https://es.whoscored.com/regions/206/tournaments/4/seasons/10803/"
        "stages/24622/refereestatistics/espa%C3%B1a-laliga-2025-2026"
    )
    print("  -> Scrapeando whoscored.com (Resultados, Chrome visible) ...")
    driver = None
    try:
        driver = build_driver(headless=False)
        driver.get(url)

        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.ID, "referee-stats-alternate")))
        time.sleep(4)

        cerrar_banners(driver)
        time.sleep(1)

        driver.execute_script(
            "document.querySelector('a[href=\"#referee-stats-alternate\"]').click();"
        )
        wait.until(lambda d: len(
            d.find_element(By.ID, "referee-stats-alternate")
             .find_elements(By.TAG_NAME, "tr")
        ) > 1)
        time.sleep(2)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        alt_div = soup.find("div", {"id": "referee-stats-alternate"})
        if not alt_div:
            return pd.DataFrame()
        tbody = alt_div.find("tbody")
        if not tbody:
            return pd.DataFrame()

        data = []
        for row in tbody.find_all("tr"):
            cells = row.find_all("td")
            if not cells or len(cells) < 5:
                continue
            a = cells[0].find("a", {"class": "tournament-link"})
            if not a:
                continue
            nombre   = nombre_desde_href(a.get("href", ""))
            ganados  = cells[2].get_text(strip=True)
            perdidos = cells[3].get_text(strip=True)
            empates  = cells[4].get_text(strip=True)
            data.append({
                "Árbitro":               nombre,
                "Ganados en casa %_ws":  ganados,
                "Perdidos en casa %_ws": perdidos,
                "Empates %_ws":          empates,
            })

        if not data:
            return pd.DataFrame()

        df = pd.DataFrame(data)
        print(f"    + {len(df)} arbitros obtenidos de whoscored.com (Resultados)")
        return df

    except Exception as e:
        print(f"    x Error en whoscored.com Resultados: {e}")
        return pd.DataFrame()
    finally:
        if driver:
            driver.quit()


# ══════════════════════════════════════════════════════════════════
# UNIFICACIÓN — hoja General
# Une las tres fuentes por nombre de árbitro normalizado.
# La fuente base es Transfermarkt (tiene el formato de nombre más limpio).
# Se hace left join con estadisticaslaliga y WhoScored.
# Los valores nulos y guiones se reemplazan por 0 para que Excel
# trate las columnas numéricas como números.
# ══════════════════════════════════════════════════════════════════
def resolver_nombre_tm(nombre, nombres_tm_norm):
    """Devuelve la clave normalizada del árbitro, aplicando NOMBRE_MAP si existe."""
    norm = normalizar(nombre)
    if norm in NOMBRE_MAP:
        return normalizar(NOMBRE_MAP[norm])
    if norm in nombres_tm_norm:
        return norm
    return norm


def unificar(df_esl, df_tm, df_ws):
    """
    Une los DataFrames de las tres fuentes por nombre de árbitro normalizado.

    Args:
        df_esl: DataFrame de estadisticaslaliga.es.
        df_tm:  DataFrame de transfermarkt.es (sirve como base del join).
        df_ws:  DataFrame de whoscored.com.

    Returns:
        DataFrame unificado, ordenado por nombre, con nulos reemplazados por 0.
    """
    nombres_tm_norm = set(df_tm["Árbitro"].apply(normalizar)) if not df_tm.empty else set()

    for df in [df_esl, df_ws]:
        if not df.empty:
            df["_key"] = df["Árbitro"].apply(
                lambda n: resolver_nombre_tm(n, nombres_tm_norm)
            )

    if not df_tm.empty:
        df_tm["_key"] = df_tm["Árbitro"].apply(normalizar)

    base = df_tm.copy() if not df_tm.empty else df_esl.copy()

    for df_other in [df_esl, df_ws]:
        if df_other.empty:
            continue
        df_other = df_other.drop(columns=["Árbitro"])
        base = pd.merge(base, df_other, on="_key", how="left")

    base = base.drop(columns=["_key"])
    cols = ["Árbitro"] + [c for c in base.columns if c != "Árbitro"]
    base = base[cols]
    base = base[base["Árbitro"].apply(normalizar) != ARBITRO_EXCLUIDO]
    base = base.sort_values("Árbitro").reset_index(drop=True)
    base = base.replace({None: 0, "-": 0})
    base = base.fillna(0)

    return base


# ══════════════════════════════════════════════════════════════════
# ESTILOS COMUNES EXCEL
# La hoja General usa dos filas de cabecera:
#   Fila 1: nombre de la fuente (colores por fuente).
#   Fila 2: nombre de cada columna.
# La hoja Partidos agrupa los partidos por árbitro con una fila de
# cabecera fusionada por árbitro antes de sus partidos.
# ══════════════════════════════════════════════════════════════════
ALT_ROW_FILL  = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
SOURCE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
COL_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=10)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Colores hoja Partidos
PARTIDOS_HEADER_COLOR = "1F3864"
PARTIDOS_LOCAL_COLOR  = "2E75B6"
PARTIDOS_VISIT_COLOR  = "ED7D31"
PARTIDOS_INFO_COLOR   = "70AD47"


def get_suffix(col_name):
    for suf in SOURCE_COLORS:
        if col_name.endswith(suf):
            return suf
    return None


# ══════════════════════════════════════════════════════════════════
# ESCRITURA EXCEL — HOJA GENERAL
# ══════════════════════════════════════════════════════════════════
def write_hoja_general(wb, df):
    """
    Escribe la hoja General con dos filas de cabecera codificadas por fuente.

    Fila 1: bloques fusionados por fuente, con el color de cada fuente.
    Fila 2: nombre legible de cada columna (de COLUMN_LABELS).
    Datos: árbitros en filas, con valores numéricos convertidos.

    Args:
        wb: Workbook de openpyxl.
        df: DataFrame unificado con todos los árbitros y columnas.
    """
    ws = wb.active
    ws.title = "General"
    ws.sheet_properties.tabColor = "1F3864"

    all_cols = list(df.columns)

    # Fila 1: cabecera "Nombre" fusionada A1:A2
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c = ws.cell(row=1, column=1, value="Nombre")
    c.font      = SOURCE_FONT
    c.fill      = PatternFill("solid", start_color=ARBITRO_COLOR, end_color=ARBITRO_COLOR)
    c.alignment = CENTER_ALIGN
    c.border    = THIN_BORDER
    ws.cell(row=2, column=1).border = THIN_BORDER

    source_labels = {
        "_esl": "estadisticaslaliga.es",
        "_tm":  "transfermarkt.es",
        "_ws":  "whoscored.com",
    }
    for sufijo, color in SOURCE_COLORS.items():
        idxs = [i + 1 for i, col in enumerate(all_cols) if col.endswith(sufijo)]
        if not idxs:
            continue
        col_start, col_end = min(idxs), max(idxs)
        fill = PatternFill("solid", start_color=color, end_color=color)
        if col_start != col_end:
            ws.merge_cells(
                start_row=1, start_column=col_start,
                end_row=1,   end_column=col_end
            )
        c = ws.cell(row=1, column=col_start, value=source_labels[sufijo])
        c.font      = SOURCE_FONT
        c.fill      = fill
        c.alignment = CENTER_ALIGN
        for ci in range(col_start, col_end + 1):
            ws.cell(row=1, column=ci).border = THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Fila 2: nombres de columnas
    for col_idx, col_name in enumerate(all_cols[1:], start=2):
        suf   = get_suffix(col_name)
        label = COLUMN_LABELS.get(col_name, col_name[: -len(suf)] if suf else col_name)
        color = SOURCE_COLORS.get(suf, ARBITRO_COLOR)
        fill  = PatternFill("solid", start_color=color, end_color=color)
        c = ws.cell(row=2, column=col_idx, value=label)
        c.font      = COL_FONT
        c.fill      = fill
        c.alignment = CENTER_ALIGN
        c.border    = THIN_BORDER

    ws.row_dimensions[2].height = 28

    # Filas de datos
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()

        for col_idx, value in enumerate(row, start=1):
            col_name = all_cols[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = BODY_FONT
            cell.fill   = fill
            cell.border = THIN_BORDER

            if col_name in COLUMNAS_TEXTO:
                cell.value     = str(value) if pd.notna(value) else ""
                cell.alignment = LEFT_ALIGN if col_name == "Árbitro" else CENTER_ALIGN
            else:
                try:
                    n = to_number(value)
                    cell.value = n
                except Exception:
                    cell.value = value
                cell.alignment = CENTER_ALIGN

    # Ancho de columnas
    for col_idx, col_name in enumerate(all_cols, start=1):
        col_letter = get_column_letter(col_idx)
        suf   = get_suffix(col_name)
        label = COLUMN_LABELS.get(col_name, col_name[: -len(suf)] if suf else col_name)
        valores = [len(str(v)) for v in df.iloc[:, col_idx - 1] if pd.notna(v)]
        max_len = max([len(label)] + valores)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════
# ESCRITURA EXCEL — HOJA PARTIDOS
# ══════════════════════════════════════════════════════════════════
def write_hoja_partidos(wb, df):
    """
    Escribe la hoja Partidos con los datos de designaciones arbitrales.

    Los partidos se agrupan por árbitro: antes de los partidos de cada árbitro
    se inserta una fila de cabecera fusionada con el nombre del árbitro en azul.
    Las columnas se colorean según si son del equipo local o visitante.

    Args:
        wb: Workbook de openpyxl.
        df: DataFrame de designaciones (una fila por partido).
    """
    ws = wb.create_sheet(title="Partidos")
    ws.sheet_properties.tabColor = "2E75B6"

    if df.empty:
        ws.cell(row=1, column=1, value="No se pudieron obtener los datos de designaciones.")
        return

    columnas = [
        "Árbitro",
        "Jornada",
        "Equipo Local",
        "Equipo Visitante",
        "Faltas Local",
        "Faltas Visitante",
        "Tarjetas Local",
        "Tarjetas Visitante",
        "F/T Local",
        "F/T Visitante",
    ]

    col_colors = {
        "Árbitro":            PARTIDOS_INFO_COLOR,
        "Jornada":            PARTIDOS_INFO_COLOR,
        "Equipo Local":       PARTIDOS_LOCAL_COLOR,
        "Equipo Visitante":   PARTIDOS_VISIT_COLOR,
        "Faltas Local":       PARTIDOS_LOCAL_COLOR,
        "Faltas Visitante":   PARTIDOS_VISIT_COLOR,
        "Tarjetas Local":     PARTIDOS_LOCAL_COLOR,
        "Tarjetas Visitante": PARTIDOS_VISIT_COLOR,
        "F/T Local":          PARTIDOS_LOCAL_COLOR,
        "F/T Visitante":      PARTIDOS_VISIT_COLOR,
    }

    # Fila 1: grupos fusionados
    # Designacion (cols 1-2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    c = ws.cell(row=1, column=1, value="Designacion")
    c.font      = SOURCE_FONT
    c.fill      = PatternFill("solid", start_color=PARTIDOS_INFO_COLOR, end_color=PARTIDOS_INFO_COLOR)
    c.alignment = CENTER_ALIGN
    c.border    = THIN_BORDER
    ws.cell(row=1, column=2).border = THIN_BORDER

    # Local (col 3)
    c = ws.cell(row=1, column=3, value="Local")
    c.font      = SOURCE_FONT
    c.fill      = PatternFill("solid", start_color=PARTIDOS_LOCAL_COLOR, end_color=PARTIDOS_LOCAL_COLOR)
    c.alignment = CENTER_ALIGN
    c.border    = THIN_BORDER

    # Visitante (col 4)
    c = ws.cell(row=1, column=4, value="Visitante")
    c.font      = SOURCE_FONT
    c.fill      = PatternFill("solid", start_color=PARTIDOS_VISIT_COLOR, end_color=PARTIDOS_VISIT_COLOR)
    c.alignment = CENTER_ALIGN
    c.border    = THIN_BORDER

    # Grupos stats: Faltas (5-6), Tarjetas (7-8), F/T (9-10)
    stat_groups = [(5, 6, "Faltas"), (7, 8, "Tarjetas"), (9, 10, "F/T")]
    for col_start, col_end, label in stat_groups:
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        c = ws.cell(row=1, column=col_start, value=label)
        c.font      = SOURCE_FONT
        c.fill      = PatternFill("solid", start_color=PARTIDOS_HEADER_COLOR, end_color=PARTIDOS_HEADER_COLOR)
        c.alignment = CENTER_ALIGN
        c.border    = THIN_BORDER
        ws.cell(row=1, column=col_end).border = THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Fila 2: nombres de columnas individuales
    for col_idx, col_name in enumerate(columnas, start=1):
        color = col_colors.get(col_name, PARTIDOS_HEADER_COLOR)
        c = ws.cell(row=2, column=col_idx, value=col_name)
        c.font      = COL_FONT
        c.fill      = PatternFill("solid", start_color=color, end_color=color)
        c.alignment = CENTER_ALIGN
        c.border    = THIN_BORDER

    ws.row_dimensions[2].height = 28

    cols_texto = {"Árbitro", "Equipo Local", "Equipo Visitante", "F/T Local", "F/T Visitante"}

    # Estilos cabecera de árbitro
    ARBITRO_HDR_FILL = PatternFill("solid", start_color=ARBITRO_COLOR, end_color=ARBITRO_COLOR)
    ARBITRO_HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    # Ordenar por árbitro alfabéticamente, luego por jornada
    df_sorted = df.sort_values(["Árbitro", "Jornada"]).reset_index(drop=True)

    current_row = 3
    alt_counter = 0

    for arbitro, grupo in df_sorted.groupby("Árbitro", sort=True):
        # Fila cabecera del árbitro (fusionada)
        n_cols = len(columnas)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=n_cols
        )
        c = ws.cell(row=current_row, column=1, value=arbitro)
        c.font      = ARBITRO_HDR_FONT
        c.fill      = ARBITRO_HDR_FILL
        c.alignment = LEFT_ALIGN
        c.border    = THIN_BORDER
        for ci in range(2, n_cols + 1):
            ws.cell(row=current_row, column=ci).border = THIN_BORDER
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Filas de partidos del árbitro
        for _, row in grupo.iterrows():
            fill = ALT_ROW_FILL if alt_counter % 2 == 0 else PatternFill()
            alt_counter += 1

            for col_idx, col_name in enumerate(columnas, start=1):
                if col_name == "Árbitro":
                    value = ""
                else:
                    value = row.get(col_name, "") if col_name in df.columns else ""

                cell = ws.cell(row=current_row, column=col_idx)
                cell.font   = BODY_FONT
                cell.fill   = fill
                cell.border = THIN_BORDER

                if col_name in cols_texto:
                    cell.value = str(value) if value else ""
                    cell.alignment = (
                        LEFT_ALIGN if col_name in {"Equipo Local", "Equipo Visitante"}
                        else CENTER_ALIGN
                    )
                else:
                    try:
                        cell.value = to_number(value)
                    except Exception:
                        cell.value = value
                    cell.alignment = CENTER_ALIGN

            current_row += 1

    # Ancho de columnas
    col_widths = {
        "Árbitro":            30,
        "Jornada":            10,
        "Equipo Local":       25,
        "Equipo Visitante":   25,
        "Faltas Local":       14,
        "Faltas Visitante":   16,
        "Tarjetas Local":     15,
        "Tarjetas Visitante": 18,
        "F/T Local":          12,
        "F/T Visitante":      14,
    }
    for col_idx, col_name in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════
# ESCRITURA DEL EXCEL COMPLETO
# ══════════════════════════════════════════════════════════════════
def write_excel(df_general, df_partidos):
    """Crea el Workbook, escribe ambas hojas y guarda el archivo."""
    wb = Workbook()
    write_hoja_general(wb, df_general)
    write_hoja_partidos(wb, df_partidos)
    wb.save(OUTPUT_FILE)
    print(f"\n Excel guardado en: {OUTPUT_FILE}")


# ══════════════════════════════════════════════════════════════════
# EJECUCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  SCRAPER ARBITROS - PRIMERA DIVISION 2025/26")
    print("=" * 60)

    print("\n[1/5] estadisticaslaliga.es - tabla arbitros")
    df_esl = scrape_estadisticaslaliga()

    print("\n[2/5] estadisticaslaliga.es - designaciones arbitrales")
    df_partidos = scrape_designaciones()

    print("\n[3/5] transfermarkt.es")
    df_tm = scrape_transfermarkt()

    print("\n[4/5] whoscored.com (Resultados)")
    print("      (Se abrira una ventana de Chrome brevemente)")
    df_ws = scrape_whoscored_resultados()

    print("\n[5/5] Unificando y escribiendo Excel ...")
    df_general = unificar(df_esl, df_tm, df_ws)
    write_excel(df_general, df_partidos)

    print(f"\n-- Arbitros en hoja General:  {len(df_general)} --")
    print(f"-- Partidos en hoja Partidos: {len(df_partidos)} --")


if __name__ == "__main__":
    main()
