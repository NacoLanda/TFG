"""
Código Duplas Peligrosas.py
===========================
Extrae en tiempo real la tabla "Asistencias al Goleador" de WhoScored para
LaLiga 2025-26 y la guarda en un Excel formateado.

Esta tabla recoge las duplas más peligrosas de la competición: para cada
goleador muestra quién le da más asistencias, cuántos goles han generado
juntos y cuántos minutos han coincidido en el campo.

El script usa Selenium en modo headless (Chrome) para cargar la página, que
requiere JavaScript para renderizar la tabla. Si WhoScored presenta un banner
de cookies, se acepta automáticamente antes de leer los datos.

Flujo:
  1. Abre Chrome headless y navega a la URL de WhoScored.
  2. Acepta el banner de cookies si aparece.
  3. Espera a que cargue el elemento player-assist-table-body.
  4. Extrae cabeceras y filas de la tabla.
  5. Escribe el Excel con formato en Datos/Duplas Peligrosas.xlsx.

Entrada:  Conexión en tiempo real a WhoScored (no hay HTML local)
Salida:   Datos/Duplas Peligrosas.xlsx

Uso: python3 "Código Duplas Peligrosas.py"
"""

import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — URL y ruta de salida
# ══════════════════════════════════════════════════════════════════
URL = (
    "https://es.whoscored.com/regions/206/tournaments/4/seasons/10803/"
    "stages/24622/playerstatistics/"
    "espa%C3%B1a-laliga-2025-2026"
)

OUTPUT_PATH = (
    "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos/Duplas Peligrosas.xlsx"
)


def crear_driver():
    """Crea y devuelve un WebDriver de Chrome en modo headless con perfil anti-detección."""
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def extraer_datos(driver):
    """
    Navega a WhoScored y extrae la tabla de asistencias al goleador.

    Maneja el banner de cookies si aparece. Espera a que cargue el
    elemento player-assist-table-body antes de extraer los datos.

    Returns:
        (cabeceras, datos): listas con los nombres de columnas y las filas.
    """
    driver.get(URL)
    wait = WebDriverWait(driver, 30)

    # Aceptar cookies si aparece el banner
    try:
        accept = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(text(),'Accept') or contains(text(),'Aceptar') or contains(text(),'Agree')]")
        ))
        accept.click()
        time.sleep(1)
    except Exception:
        pass

    # Esperar a que cargue la tabla de asistencias al goleador
    wait.until(EC.presence_of_element_located((By.ID, "player-assist-table-body")))
    time.sleep(2)

    # Extraer cabeceras
    thead = driver.find_element(By.ID, "player-assist-table-head")
    cabeceras = [th.text.strip() for th in thead.find_elements(By.TAG_NAME, "th")]

    # Extraer filas
    tbody = driver.find_element(By.ID, "player-assist-table-body")
    filas = tbody.find_elements(By.TAG_NAME, "tr")

    datos = []
    for fila in filas:
        celdas = fila.find_elements(By.TAG_NAME, "td")
        fila_datos = [celda.text.strip() for celda in celdas]
        if fila_datos:
            datos.append(fila_datos)

    return cabeceras, datos


def guardar_excel(cabeceras, datos):
    """
    Genera el Excel con las duplas peligrosas (goleador + asistidor).

    Aplica formato con cabecera oscura, filas alternadas y ajuste
    automático de ancho. Guarda en la ruta definida en OUTPUT_PATH.
    """
    COLOR_HEADER_BG  = "1A3A5C"
    COLOR_HEADER_FG  = "FFFFFF"
    COLOR_FILA_PAR   = "DCE6F1"
    COLOR_FILA_IMPAR = "FFFFFF"
    COLOR_BORDE      = "A6BEDB"

    borde = Border(
        left=Side(style="thin", color=COLOR_BORDE),
        right=Side(style="thin", color=COLOR_BORDE),
        top=Side(style="thin", color=COLOR_BORDE),
        bottom=Side(style="thin", color=COLOR_BORDE),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias al Goleador"

    # Cabecera
    ws.append(cabeceras)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[1].height = 28

    # Datos
    for fila in datos:
        ws.append(fila)
        row_idx = ws.max_row
        color = COLOR_FILA_PAR if row_idx % 2 == 0 else COLOR_FILA_IMPAR
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.border = borde
            if col_idx in (1, len(cabeceras)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_idx].height = 20

    # Ancho de columnas automático
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_PATH)


def main():
    print("Conectando con WhoScored...")
    driver = crear_driver()
    try:
        cabeceras, datos = extraer_datos(driver)
    finally:
        driver.quit()

    if not datos:
        print("No se pudieron extraer datos. Comprueba tu conexión o si WhoScored ha cambiado su estructura.")
        return

    print(f"Extraídas {len(datos)} filas. Generando Excel...")
    guardar_excel(cabeceras, datos)
    print(f"Archivo guardado en: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
