"""
Código Lluvias.py
=================
Descarga datos de precipitación histórica y por partido de la API
Open-Meteo y los consolida en un Excel actualizable.

El script es incremental: si Datos Lluvias.xlsx ya existe, solo descarga
los datos desde el día siguiente al último registrado. Si no existe, parte
de DEFAULT_START (inicio de la temporada).

Flujo:
  1. Determina el rango de fechas a descargar (get_start_iso).
  2. Descarga precipitación diaria acumulada por estadio (fetch_daily).
  3. Lee Partidos.xlsx para saber qué partidos se jugaron en ese rango.
  4. Descarga datos horarios para cada partido (fetch_hourly_for_date),
     identificando los mm caídos durante las horas del encuentro.
  5. Lee los datos previos del Excel existente (read_existing_data).
  6. Construye el Excel nuevo desde cero combinando datos previos y nuevos
     (nunca modifica el archivo original, siempre lo reescribe limpio).

Salidas en Datos/Datos Lluvias.xlsx:
  - Hoja "Resumen":       precipitación total y partidos con lluvia por equipo
  - Hoja "Datos diarios": registro partido a partido con lluvia horaria

Uso: python3 "Código Lluvias.py"
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

# ══════════════════════════════════════════════════════════════════
# DATOS DE ESTADIOS
# Coordenadas GPS necesarias para consultar la API de Open-Meteo.
# COVERED_STADIUMS: estadios con techo retráctil (no se consultan).
# ══════════════════════════════════════════════════════════════════
STADIUMS = [
    {"club": "Barcelona",        "estadio": "Spotify Camp Nou",               "lat": 41.3809, "lon":  2.1228},
    {"club": "Real Madrid",      "estadio": "Santiago Bernabéu",              "lat": 40.4531, "lon": -3.6883},
    {"club": "Atlético Madrid",  "estadio": "Cívitas Metropolitano",          "lat": 40.4361, "lon": -3.5995},
    {"club": "Villarreal",       "estadio": "Estadio de la Cerámica",         "lat": 39.9441, "lon": -0.1036},
    {"club": "Real Betis",       "estadio": "Estadio Benito Villamarín",      "lat": 37.3561, "lon": -5.9817},
    {"club": "Celta Vigo",       "estadio": "Abanca Balaídos",                "lat": 42.2117, "lon": -8.7395},
    {"club": "Athletic Club",    "estadio": "San Mamés",                      "lat": 43.2642, "lon": -2.9496},
    {"club": "Real Sociedad",    "estadio": "Reale Arena",                    "lat": 43.3015, "lon": -1.9731},
    {"club": "Girona",           "estadio": "Estadio Municipal de Montilivi", "lat": 41.9609, "lon":  2.8306},
    {"club": "Osasuna",          "estadio": "El Sadar",                       "lat": 42.7966, "lon": -1.6372},
    {"club": "Getafe",           "estadio": "Coliseum Alfonso Pérez",         "lat": 40.3259, "lon": -3.7172},
    {"club": "Rayo Vallecano",   "estadio": "Estadio de Vallecas",            "lat": 40.3919, "lon": -3.6601},
    {"club": "Mallorca",         "estadio": "Estadi de Son Moix",             "lat": 39.5898, "lon":  2.6653},
    {"club": "Alavés",           "estadio": "Mendizorrotza",                  "lat": 42.8418, "lon": -2.6833},
    {"club": "Espanyol",         "estadio": "RCDE Stadium",                   "lat": 41.3478, "lon":  2.0750},
    {"club": "Sevilla",          "estadio": "Ramón Sánchez-Pizjuán",          "lat": 37.3838, "lon": -5.9706},
    {"club": "Valencia",         "estadio": "Mestalla",                       "lat": 39.4747, "lon": -0.3583},
    {"club": "Levante",          "estadio": "Estadio Ciudad de Valencia",     "lat": 39.4936, "lon": -0.3580},
    {"club": "Elche",            "estadio": "Estadio Martínez Valero",        "lat": 38.2669, "lon": -0.6958},
    {"club": "Oviedo",           "estadio": "Estadio Carlos Tartiere",        "lat": 43.3524, "lon": -5.8731},
]

COVERED_STADIUMS = {"Real Madrid"}
CLUB_TO_STADIUM  = {s["club"]: s for s in STADIUMS}

END_ISO       = date.today().strftime("%Y-%m-%d")
BASE_PATH     = "/Users/NacoLG/Documentos/UFV 4/TFG/Ingeniería del Dato/Datos"
OUTPUT_PATH   = os.path.join(BASE_PATH, "Datos Lluvias.xlsx")
PARTIDOS_PATH = os.path.join(BASE_PATH, "Partidos.xlsx")
DEFAULT_START = "2025-08-15"


def fmt_date(iso):
    """Convierte fecha ISO 'YYYY-MM-DD' al formato legible 'DD-MM-YYYY'."""
    y, m, d = iso.split("-")
    return f"{d}-{m}-{y}"


def get_start_iso():
    """
    Devuelve la fecha desde la que hay que descargar datos nuevos:
    - Si existe Datos Lluvias.xlsx, lee la última fecha en "Datos diarios" y usa el día siguiente.
    - Si no existe, usa DEFAULT_START.
    """
    if not os.path.exists(OUTPUT_PATH):
        print(f"  → Excel no encontrado. Usando fecha de inicio por defecto: {fmt_date(DEFAULT_START)}.")
        return DEFAULT_START
    try:
        wb = openpyxl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Datos diarios" not in wb.sheetnames:
            wb.close()
            print(f"  → Hoja 'Datos diarios' no encontrada. Usando fecha de inicio por defecto: {fmt_date(DEFAULT_START)}.")
            return DEFAULT_START
        ws = wb["Datos diarios"]
        last_date_fmt = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                last_date_fmt = str(row[0]).strip()  # formato DD-MM-YYYY
        wb.close()
        if last_date_fmt:
            # Convertir DD-MM-YYYY a YYYY-MM-DD
            d, m, y = last_date_fmt.split("-")
            last_iso = f"{y}-{m}-{d}"
            # El día siguiente al último registrado
            from datetime import datetime, timedelta
            next_day = (datetime.strptime(last_iso, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            print(f"  → Último dato en Datos Lluvias.xlsx: {last_date_fmt}. Descargando desde: {fmt_date(next_day)}.")
            return next_day
    except Exception as e:
        print(f"  → No se pudo leer Datos Lluvias.xlsx: {e}")
    print(f"  → Sin datos previos. Usando fecha de inicio por defecto: {fmt_date(DEFAULT_START)}.")
    return DEFAULT_START


START_ISO   = get_start_iso()
START_LABEL = fmt_date(START_ISO)
END_LABEL   = fmt_date(END_ISO)


def parse_lluvia(value):
    """
    Interpreta el valor de lluvia almacenado en Partidos.xlsx.

    Devuelve:
        None  → estadio cubierto (no computa)
        0.0   → no llovió
        float → milímetros de lluvia durante el partido
    """
    if value is None:
        return 0.0
    s = str(value).strip()
    if s == "Estadio Cubierto":
        return None
    if s == "No Llovió":
        return 0.0
    try:
        return float(s.replace("mm", "").replace(",", ".").strip())
    except ValueError:
        return 0.0


def fetch_daily(s):
    """
    Descarga precipitación diaria acumulada para un estadio en el rango de fechas.

    Usa la API histórica de Open-Meteo con las coordenadas del estadio.

    Returns:
        (total_mm, rainy_days): total de mm y número de días con > 0.1 mm.
    """
    url = (
        "https://archive-api.open-meteo.com/v1/archive"
        f"?latitude={s['lat']}&longitude={s['lon']}"
        f"&start_date={START_ISO}&end_date={END_ISO}"
        "&daily=precipitation_sum&timezone=Europe%2FMadrid"
    )
    data = requests.get(url, timeout=30).json()
    precip = [v if v is not None else 0.0 for v in data["daily"]["precipitation_sum"]]
    dates  = data["daily"]["time"]
    total      = round(sum(precip), 1)
    rainy_days = sum(1 for v in precip if v > 0.1)
    return total, rainy_days


def fetch_hourly_for_date(s, date_iso):
    """
    Descarga precipitación hora a hora para un estadio en una fecha concreta.

    Returns:
        Lista de tuplas (hora_HH:MM, mm) solo para las horas con lluvia > 0.
    """
    url = (
        "https://archive-api.open-meteo.com/v1/archive"
        f"?latitude={s['lat']}&longitude={s['lon']}"
        f"&start_date={date_iso}&end_date={date_iso}"
        "&hourly=precipitation&timezone=Europe%2FMadrid"
    )
    data = requests.get(url, timeout=30).json()
    times  = data["hourly"]["time"]
    precip = data["hourly"]["precipitation"]
    return [
        (t.split("T")[1], round(p, 1))
        for t, p in zip(times, precip)
        if p is not None and p > 0
    ]


def load_matches_ordered():
    """
    Lee Partidos.xlsx y extrae los partidos jugados en el rango de fechas.

    También acumula totales de lluvia por equipo (local y visitante)
    para rellenar la hoja Resumen del Excel de salida.

    Returns:
        matches:    Lista de (fecha_iso, club_local) de los partidos en rango.
        rain_local: Dict equipo → {total mm, partidos con lluvia} como local.
        rain_away:  Dict equipo → {total mm, partidos con lluvia} como visitante.
    """
    wb = openpyxl.load_workbook(PARTIDOS_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers    = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    date_col   = headers.index("Date")
    home_col   = headers.index("Home")
    away_col   = headers.index("Away")
    lluvia_col = headers.index("Lluvia")

    matches    = []
    rain_local = {s["club"]: {"total": 0.0, "con_lluvia": 0} for s in STADIUMS}
    rain_away  = {s["club"]: {"total": 0.0, "con_lluvia": 0} for s in STADIUMS}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[date_col]
        home     = row[home_col]
        away     = row[away_col]
        lluvia   = row[lluvia_col]

        if raw_date is None or home is None:
            continue
        date_iso = raw_date.strftime("%Y-%m-%d") if hasattr(raw_date, "strftime") else str(raw_date).strip()[:10]

        if date_iso < START_ISO or date_iso > END_ISO:
            continue
        if home not in CLUB_TO_STADIUM:
            continue

        matches.append((date_iso, home))
        mm = parse_lluvia(lluvia)

        if home not in COVERED_STADIUMS and mm is not None:
            rain_local[home]["total"]      = round(rain_local[home]["total"] + mm, 2)
            if mm > 0:
                rain_local[home]["con_lluvia"] += 1

        if away in rain_away and mm is not None:
            rain_away[away]["total"]      = round(rain_away[away]["total"] + mm, 2)
            if mm > 0:
                rain_away[away]["con_lluvia"] += 1

    wb.close()
    return matches, rain_local, rain_away


def read_existing_data():
    """
    Lee el Excel existente y devuelve:
    - resumen_prev: dict club -> {total, rainy, rl_total, rl_lluvia, ra_total, ra_lluvia}
    - diarios_prev: lista de tuplas (fecha_fmt, estadio, club, horas_str) ya guardadas
    """
    resumen_prev = {s["club"]: {"total": 0.0, "rainy": 0, "rl_total": 0.0,
                                "rl_lluvia": 0, "ra_total": 0.0, "ra_lluvia": 0}
                   for s in STADIUMS}
    diarios_prev = []

    if not os.path.exists(OUTPUT_PATH):
        return resumen_prev, diarios_prev

    try:
        wb = openpyxl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)

        # Leer Resumen
        if "Resumen" in wb.sheetnames:
            ws = wb["Resumen"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                club = row[0]
                if club is None or club == "PROMEDIO" or club not in CLUB_TO_STADIUM:
                    continue
                def safe(v):
                    if v is None or isinstance(v, str):
                        return 0.0
                    return float(v)
                resumen_prev[club] = {
                    "total":     safe(row[2]),
                    "rainy":     int(safe(row[3])),
                    "rl_total":  safe(row[4]) if club not in COVERED_STADIUMS else 0.0,
                    "rl_lluvia": int(safe(row[5])) if club not in COVERED_STADIUMS else 0,
                    "ra_total":  safe(row[6]),
                    "ra_lluvia": int(safe(row[7])),
                }

        # Leer Datos diarios
        if "Datos diarios" in wb.sheetnames:
            ws2 = wb["Datos diarios"]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                diarios_prev.append((
                    str(row[0]).strip(),  # fecha formateada DD-MM-YYYY
                    str(row[1]).strip() if row[1] else "",
                    str(row[2]).strip() if row[2] else "",
                    str(row[3]).strip() if row[3] else "",
                ))

        wb.close()
        print(f"  → Excel existente leído: {len(diarios_prev)} filas en Datos diarios.")
    except Exception as e:
        print(f"  → No se pudo leer el Excel existente: {e}. Se creará desde cero.")

    return resumen_prev, diarios_prev


def make_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr_cell(cell, value, bg="1A5276", align="center"):
    """Aplica estilo de cabecera (fondo azul, texto blanco en negrita) a una celda."""
    cell.value     = value
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = make_border()


def data_cell(cell, value, fill_color, align="center", number_format=None):
    """Aplica valor, fuente Arial y relleno de color a una celda de datos."""
    cell.value     = value
    cell.font      = Font(name="Arial")
    cell.fill      = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = make_border()
    if number_format:
        cell.number_format = number_format


def apply_border_to_merged(ws, cell_range, fill_color):
    """Añade borde solo en el perímetro exterior de un rango de celdas fusionadas."""
    thin = Side(style="thin")
    rows = list(ws[cell_range])
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            top    = thin if r_idx == 0 else Side(style=None)
            bottom = thin if r_idx == len(rows) - 1 else Side(style=None)
            left   = thin if c_idx == 0 else Side(style=None)
            right  = thin if c_idx == len(row) - 1 else Side(style=None)
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            cell.fill   = PatternFill("solid", fgColor=fill_color)


def build_excel(new_results, new_matches, hourly_data, new_rain_local, new_rain_away,
                prev_resumen, prev_diarios):
    """
    Construye el Excel limpio desde cero combinando datos previos y nuevos.

    Nunca modifica el archivo existente: siempre genera un libro nuevo y
    lo sobreescribe completo, garantizando que el resultado es consistente.
    Los datos de resumen se suman a los previos antes de escribir.

    Args:
        new_results:    Dict {club: (total_mm, dias_lluvia)} de esta ejecución.
        new_matches:    Lista de (fecha_iso, club_local) de partidos en el rango.
        hourly_data:    Dict {(club, fecha_iso): [(hora, mm), ...]} de esta ejecución.
        new_rain_local: Dict {club: {total, con_lluvia}} como local de esta ejecución.
        new_rain_away:  Dict {club: {total, con_lluvia}} como visitante de esta ejecución.
        prev_resumen:   Dict {club: {total, rainy, ...}} del Excel existente.
        prev_diarios:   Lista de tuplas del Excel existente (ya guardadas).

    Returns:
        Workbook de openpyxl listo para guardar.
    """
    wb = openpyxl.Workbook()
    ALT       = "D6EAF8"
    EVEN      = "FFFFFF"
    SUM_COLOR = "AED6F1"

    # ── Combinar datos de Resumen ────────────────────────────────────────────
    combined = {}
    for s in STADIUMS:
        club = s["club"]
        prev = prev_resumen[club]
        nr   = new_results.get(club, (0.0, 0))
        rl   = new_rain_local.get(club, {"total": 0.0, "con_lluvia": 0})
        ra   = new_rain_away.get(club,  {"total": 0.0, "con_lluvia": 0})
        combined[club] = {
            "total":     round(prev["total"]     + nr[0],              1),
            "rainy":     int(prev["rainy"]        + nr[1]),
            "rl_total":  round(prev["rl_total"]  + rl["total"],        1),
            "rl_lluvia": int(prev["rl_lluvia"]   + rl["con_lluvia"]),
            "ra_total":  round(prev["ra_total"]  + ra["total"],        1),
            "ra_lluvia": int(prev["ra_lluvia"]   + ra["con_lluvia"]),
        }

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"

    ws.merge_cells("A1:I1")
    ws["A1"].value     = f"Precipitación acumulada · Estadios LaLiga EA Sports 2025-26 · {fmt_date(DEFAULT_START)} – {END_LABEL}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A5276")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers_res = [
        "Club", "Estadio", "Total (mm)", "Días con lluvia",
        "Total (mm) como Local", "Partidos con lluvia como Local",
        "Total (mm) como Visitante", "Partidos con lluvia como Visitante",
        "Coordenadas"
    ]
    for col, h in enumerate(headers_res, 1):
        hdr_cell(ws.cell(row=2, column=col), h, bg="2980B9")

    sorted_clubs = sorted(
        STADIUMS,
        key=lambda s: combined[s["club"]]["total"],
        reverse=True
    )
    covered_rows = []

    for i, s in enumerate(sorted_clubs):
        row  = i + 3
        fill = ALT if i % 2 == 0 else EVEN
        club = s["club"]
        c    = combined[club]

        data_cell(ws.cell(row=row, column=1), club,         fill, align="left")
        data_cell(ws.cell(row=row, column=2), s["estadio"],  fill, align="left")
        data_cell(ws.cell(row=row, column=3), c["total"],    fill, align="center", number_format="#,##0.0")
        data_cell(ws.cell(row=row, column=4), c["rainy"],    fill, align="center")

        if club in COVERED_STADIUMS:
            covered_rows.append(row)
            merge_range = f"E{row}:F{row}"
            ws.merge_cells(merge_range)
            merged_cell = ws.cell(row=row, column=5)
            merged_cell.value     = "Estadio Cubierto"
            merged_cell.font      = Font(italic=True, color="5D6D7E", name="Arial")
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            apply_border_to_merged(ws, merge_range, fill)
        else:
            data_cell(ws.cell(row=row, column=5), c["rl_total"],  fill, align="center", number_format="#,##0.0")
            data_cell(ws.cell(row=row, column=6), c["rl_lluvia"], fill, align="center")

        data_cell(ws.cell(row=row, column=7), c["ra_total"],  fill, align="center", number_format="#,##0.0")
        data_cell(ws.cell(row=row, column=8), c["ra_lluvia"], fill, align="center")
        data_cell(ws.cell(row=row, column=9), f"{s['lat']}, {s['lon']}", fill, align="center")

    last_data_row = len(STADIUMS) + 2
    summary_row   = last_data_row + 2
    all_rows      = list(range(3, last_data_row + 1))
    open_rows     = [r for r in all_rows if r not in covered_rows]

    def avg_all(col_letter):
        return f"=AVERAGE({','.join(f'{col_letter}{r}' for r in all_rows)})"

    def avg_open(col_letter):
        return f"=AVERAGE({','.join(f'{col_letter}{r}' for r in open_rows)})"

    ws.cell(row=summary_row, column=1).value = "PROMEDIO"
    for col, nf in [(3, "#,##0.0"), (4, "0.0")]:
        c = ws.cell(row=summary_row, column=col)
        c.value, c.number_format = avg_all(get_column_letter(col)), nf
    for col, nf in [(5, "#,##0.0"), (6, "0.0")]:
        c = ws.cell(row=summary_row, column=col)
        c.value, c.number_format = avg_open(get_column_letter(col)), nf
    for col, nf in [(7, "#,##0.0"), (8, "0.0")]:
        c = ws.cell(row=summary_row, column=col)
        c.value, c.number_format = avg_all(get_column_letter(col)), nf

    for col in range(1, 10):
        c = ws.cell(row=summary_row, column=col)
        c.fill      = PatternFill("solid", fgColor=SUM_COLOR)
        c.font      = Font(bold=True, name="Arial")
        c.border    = make_border()
        c.alignment = Alignment(horizontal="left" if col <= 2 else "center", vertical="center")

    for col, w in zip("ABCDEFGHI", [24, 30, 14, 17, 22, 26, 24, 28, 22]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    # ── Hoja 2: Datos diarios ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Datos diarios")
    col_headers = ["Fecha", "Estadio", "Club local", "Horas con lluvia (hora: mm)"]
    for col, h in enumerate(col_headers, 1):
        hdr_cell(ws2.cell(row=1, column=col), h, bg="1A5276")

    # Construir conjunto de claves existentes para evitar duplicados
    existing_keys = {(r[0], r[2]) for r in prev_diarios}

    # Primero escribir las filas previas
    all_diario_rows = list(prev_diarios)

    # Añadir las filas nuevas que no sean duplicados
    new_added = 0
    for date_iso, club in new_matches:
        fecha_fmt = fmt_date(date_iso)
        if (fecha_fmt, club) in existing_keys:
            continue
        s = CLUB_TO_STADIUM[club]
        if club in COVERED_STADIUMS:
            horas_str = "Estadio Cubierto"
        else:
            horas = hourly_data.get((club, date_iso), [])
            horas_str = "  |  ".join(f"{h}: {mm} mm" for h, mm in horas) if horas else "Sin lluvia"
        all_diario_rows.append((fecha_fmt, s["estadio"], club, horas_str))
        existing_keys.add((fecha_fmt, club))
        new_added += 1

    # Escribir todas las filas
    for row_idx, (fecha, estadio, club, horas_str) in enumerate(all_diario_rows, 2):
        fill = ALT if row_idx % 2 == 0 else EVEN
        for col, (val, al, nf) in enumerate([
            (fecha,     "center", None),
            (estadio,   "left",   None),
            (club,      "left",   None),
            (horas_str, "left",   None),
        ], 1):
            data_cell(ws2.cell(row=row_idx, column=col), val, fill, align=al, number_format=nf)

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 80
    ws2.freeze_panes = "A2"

    print(f"  → Datos diarios: {new_added} filas nuevas añadidas ({len(all_diario_rows)} filas en total).")
    return wb


def main():
    """
    Punto de entrada principal del script.

    Orquesta la descarga, combinación con datos previos y escritura del Excel
    en el orden correcto: diarios → partidos → horarios → previos → Excel.
    """
    print(f"\nRango de actualización: {START_LABEL} → {END_LABEL}\n")

    print("Descargando datos diarios de Open-Meteo...\n")
    new_results = {}
    for s in STADIUMS:
        print(f"  · {s['club']:20s}  [{s['estadio']}]...", end=" ", flush=True)
        total, rainy_days = fetch_daily(s)
        new_results[s["club"]] = (total, rainy_days)
        print(f"{total:,.1f} mm  |  {rainy_days} días de lluvia")

    print("\nLeyendo Partidos.xlsx...")
    matches, rain_local, rain_away = load_matches_ordered()
    print(f"  → {len(matches)} partidos encontrados en el rango de fechas.")

    print("\nDescargando datos horarios para cada partido...\n")
    hourly_data = {}
    for date_iso, club in matches:
        if club in COVERED_STADIUMS:
            continue
        if (club, date_iso) in hourly_data:
            continue
        s = CLUB_TO_STADIUM[club]
        print(f"  · {club:20s}  {fmt_date(date_iso)}...", end=" ", flush=True)
        horas = fetch_hourly_for_date(s, date_iso)
        hourly_data[(club, date_iso)] = horas
        print(f"{len(horas)} horas con lluvia" if horas else "sin lluvia")

    print("\nLeyendo datos existentes del Excel...")
    prev_resumen, prev_diarios = read_existing_data()

    print("\nGenerando Excel limpio...")
    os.makedirs(BASE_PATH, exist_ok=True)
    wb = build_excel(new_results, matches, hourly_data, rain_local, rain_away,
                     prev_resumen, prev_diarios)
    wb.save(OUTPUT_PATH)
    print(f"\n✅ Archivo guardado en:\n   {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
